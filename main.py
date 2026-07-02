from fastapi import FastAPI, UploadFile, File, Form, Request, HTTPException
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import xlrd, openpyxl, shutil, tempfile, os, json, re, math, zipfile


# ============================================================
# 入力Excel読み取り：xls / xlsx / xlsm を同一インターフェースで開く
# 既存コードは xlrd の cell_value(r,c)（0始まり）・nrows・ncols・
# sheet_by_name・sheets() を使うため、xlsx/xlsm でも同じ形で返す
# ============================================================
class _OpxSheet:
    def __init__(self, ws):
        self.name = ws.title
        self.nrows = ws.max_row or 0
        self.ncols = ws.max_column or 0
        self._cells = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    self._cells[(cell.row - 1, cell.column - 1)] = cell.value

    def cell_value(self, r, c):
        v = self._cells.get((r, c), "")
        return "" if v is None else v


class _OpxBook:
    def __init__(self, path):
        wb = openpyxl.load_workbook(path, data_only=True)
        self._names = list(wb.sheetnames)
        self._sheets = {n: _OpxSheet(wb[n]) for n in self._names}

    def sheet_by_name(self, name):
        return self._sheets[name]

    def sheet_names(self):
        return list(self._names)

    def sheets(self):
        return [self._sheets[n] for n in self._names]


def open_workbook_any(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return _OpxBook(path)
    return xlrd.open_workbook(path)


def _upload_suffix(file, default=".xls"):
    """アップロードファイル名から拡張子を取得（.xls/.xlsx/.xlsm のみ許可、それ以外はdefault）"""
    name = getattr(file, "filename", "") or ""
    ext = os.path.splitext(name)[1].lower()
    return ext if ext in (".xls", ".xlsx", ".xlsm") else default


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "テンプレ", "（橋梁）03_データ記録様式(R6)_260601.xlsx")
INSPECTION_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "テンプレ", "（橋梁）01_点検記録様式(R6)_260601.xlsx")
CHOSHO_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "テンプレ", "道路橋記録様式_R6_3_.xlsx")

def _xls_str(ws, r, c):
    """セル値を文字列として取得（floatは整数文字列化）"""
    v = ws.cell_value(r, c)
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v)
    return str(v).strip() if v not in (None, "") else ""


def _xls_int(ws, r, c, default=0):
    """セル値を整数として取得（変換不可ならdefault）"""
    v = ws.cell_value(r, c)
    try:
        return int(v)
    except (ValueError, TypeError):
        return default


def read_excel(path: str) -> dict:
    wb = open_workbook_any(path)
    ws = wb.sheet_by_name("その１")

    # 調書更新年月日（年・月・日を結合）
    y = _xls_int(ws, 6, 35)
    m = _xls_int(ws, 6, 37)
    d = _xls_int(ws, 6, 39)
    chosho_date = f"{y}年{m}月{d}日" if y else ""

    # 下部構造形式・基礎形式（橋台・橋脚を結合）
    keishiki_kabu = _xls_str(ws, 15, 6)
    keishiki_kakyaku = _xls_str(ws, 16, 6)
    kaso_kabu = _xls_str(ws, 15, 27)
    kaso_kakyaku = _xls_str(ws, 16, 27)
    kabu_keishiki = "\n".join(s for s in [
        f"橋台：{keishiki_kabu}" if keishiki_kabu else "",
        f"橋脚：{keishiki_kakyaku}" if keishiki_kakyaku else "",
    ] if s)
    kaso_keishiki = "\n".join(s for s in [
        f"橋台：{kaso_kabu}" if kaso_kabu else "",
        f"橋脚：{kaso_kakyaku}" if kaso_kakyaku else "",
    ] if s)

    return {
        "橋梁名":     ws.cell_value(5, 3),
        "フリガナ":   ws.cell_value(4, 3),
        "路線名":     ws.cell_value(4, 15),
        "管理者":     ws.cell_value(4, 24),
        "所在地":     ws.cell_value(6, 4),
        "橋長":       ws.cell_value(11, 14),
        "径間数":     int(ws.cell_value(11, 21)),
        "上部構造形式": ws.cell_value(13, 4),
        "完成年号":   ws.cell_value(11, 4),
        "完成年":     int(ws.cell_value(11, 5)),
        "完成月":     int(ws.cell_value(11, 7)),
        # --- 点検記録様式その1用 追加項目 ---
        "所在地至":     _xls_str(ws, 8, 4),
        "起点側緯度":   _xls_str(ws, 1, 23),
        "起点側経度":   _xls_str(ws, 2, 23),
        "終点側緯度":   _xls_str(ws, 1, 30),
        "終点側経度":   _xls_str(ws, 2, 30),
        "施設ID":       _xls_str(ws, 1, 36),
        "距離標自":     _xls_str(ws, 6, 16),
        "距離標至":     _xls_str(ws, 8, 16),
        "調書更新年月日": chosho_date,
        "適用示方書":   _xls_str(ws, 17, 4),
        "全幅員":       _xls_str(ws, 17, 13),
        "有効幅員":     _xls_str(ws, 19, 13),
        "下部構造形式": kabu_keishiki,
        "基礎形式":     kaso_keishiki,
    }


def write_sheet_header(ws, data: dict, span_no: int, write_span_no: bool = True):
    """ヘッダー情報をシートに書き込む共通処理"""
    ws["H5"]  = data["フリガナ"]
    ws["H6"]  = data["橋梁名"]
    ws["AE5"] = data["路線名"]
    ws["AQ5"] = data["管理者"]
    if write_span_no:
        ws["AE2"] = span_no  # 径間番号

def copy_row_format(ws, src_row: int, dst_row: int):
    """
    src_rowの書式（罫線・塗り・フォント・配置・結合）をdst_rowにコピーする。
    """
    from copy import copy
    from openpyxl.utils import get_column_letter

    # 既存のdst_row結合を解除
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= dst_row <= merged.max_row:
            ws.unmerge_cells(str(merged))

    # 書式コピー（結合解除後・結合複製前に実施 → 全セルにアクセス可能）
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if type(dst).__name__ == "MergedCell":
            continue
        if src.has_style:
            dst.font          = copy(src.font)
            dst.border        = copy(src.border)
            dst.fill          = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment     = copy(src.alignment)
        dst.value = None

    # 結合をsrc_rowからdst_rowに複製（書式コピー後）
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row == src_row == merged.max_row:
            new_range = f"{get_column_letter(merged.min_col)}{dst_row}:{get_column_letter(merged.max_col)}{dst_row}"
            ws.merge_cells(new_range)

    # 行の高さをコピー
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def safe_write(ws, row, col, value):
    """結合セルの左上セルのみ書き込む"""
    c = ws.cell(row, col)
    if type(c).__name__ != "MergedCell":
        c.value = value


def expand_members(members: list) -> list:
    """
    部材リストを要素番号ごとに1行ずつに展開する。
    element_noが「0101, 0102」の場合 → 2行に分割。
    isDeleted=Trueの部材はスキップ。
    """
    rows = []
    for m in members:
        if m.get("isDeleted"):
            continue
        elem_nos = [e.strip() for e in m.get("element_no", "").split(",") if e.strip()] or [""]
        for en in elem_nos:
            rows.append({**m, "element_no": en})
    return rows


def write_members(ws, members: list, damages: list, span_no: int, start_row: int = 12):
    """
    部材リストと損傷データをシートに書き込む。
    damages: DamageListのdamagesData（span_noでフィルタ済み想定）
    行数が不足する場合はテンプレート行をコピーして追加する。
    """
    expanded = expand_members(members)

    # 損傷データをelement_noをキーにしてマッピング
    # {element_no: [damage, ...]}
    dmg_map = {}
    for d in damages:
        if d.get("isNote"):
            continue
        elem_current = d.get("elementNoCurrent", "")
        elem_keys = [e.strip() for e in elem_current.split(",") if e.strip()] if "," in (elem_current or "") else [elem_current]
        for ek in elem_keys:
            key = d.get("symbol", "") + "|" + ek
            if key not in dmg_map:
                dmg_map[key] = []
            dmg_map[key].append(d)
        # subDamagesの別部材も独立キーで登録（※上記以外の損傷の別部材を3-3に反映）
        for sub in d.get("subDamages", []):
            sub_sym = sub.get("symbol", "")
            sub_elem = sub.get("elementNoCurrent", "")
            if not sub_sym or sub_sym == d.get("symbol", ""):
                continue
            sub_key = sub_sym + "|" + sub_elem
            sub_entry = {
                "symbol": sub_sym,
                "elementNoCurrent": sub_elem,
                "memberName": sub.get("memberName", ""),
                "dmgType": sub.get("dmgType", ""),
                "currDeg": sub.get("currDeg", ""),
                "prevDeg": sub.get("prevDeg", ""),
                "detail": sub.get("detail", ""),
                "pattern": sub.get("pattern", ""),
                "bunrui": sub.get("bunrui", ""),
                "bunruiText": sub.get("bunruiText", ""),
                "subDamages": [],
                "isNote": False,
            }
            if sub_key not in dmg_map:
                dmg_map[sub_key] = []
            dmg_map[sub_key].append(sub_entry)

    # 実際に書き込む行リストを生成
    write_rows = []
    for m in expanded:
        en = m.get("element_no", "")
        symbol = m.get("symbol", "")
        zairyo = m.get("zairyo", "")
        dmg_list = dmg_map.get(symbol + "|" + en, [])

        targets = get_target_damages(symbol, zairyo)

        if dmg_list:
            # 実損傷をDXF番号キーでマップ（②重複回避：同一丸数字は最初の1件、パターンは全件収集）
            dxf_map = {}
            dxf_patterns = {}  # ③パターン複数収集用
            for d in dmg_list:
                key = circled_char(d.get("dmgType", ""))
                if key:
                    if key not in dxf_map:
                        dxf_map[key] = d
                    pat = d.get("pattern", "")
                    if pat:
                        dxf_patterns.setdefault(key, [])
                        if pat not in dxf_patterns[key]:
                            dxf_patterns[key].append(pat)
                # ①subDamagesも全件登録
                for sub in d.get("subDamages", []):
                    sk = circled_char(sub.get("dmgType", ""))
                    if sk:
                        if sk not in dxf_map:
                            dxf_map[sk] = {
                                "dmgType": sub.get("dmgType",""),
                                "currDeg": sub.get("currDeg",""),
                                "prevDeg": sub.get("prevDeg",""),
                                "detail": sub.get("detail",""),
                                "pattern": sub.get("pattern",""),
                                "bunrui": sub.get("bunrui",""),
                                "bunruiText": sub.get("bunruiText",""),
                            }
                        sp = sub.get("pattern", "")
                        if sp:
                            dxf_patterns.setdefault(sk, [])
                            if sp not in dxf_patterns[sk]:
                                dxf_patterns[sk].append(sp)
            if targets:
                outputted = set()
                for tgt in targets:
                    tgt_key = circled_char(tgt)
                    if tgt_key in dxf_map:
                        write_rows.append((m, dxf_map[tgt_key], dxf_patterns.get(tgt_key, [])))
                        outputted.add(tgt_key)
                    else:
                        write_rows.append((m, {"dmgType": tgt, "currDeg": "a", "prevDeg": "", "detail": "", "pattern": "", "bunrui": "", "bunruiText": ""}, []))
                for key, d in dxf_map.items():
                    if key not in outputted:
                        write_rows.append((m, d, dxf_patterns.get(key, [])))
            else:
                # targetsなしの場合もdxf_map経由で重複除去して出力
                for key, d in dxf_map.items():
                    write_rows.append((m, d, dxf_patterns.get(key, [])))
        else:
            if targets:
                # 損傷なし → 全種類NON/a
                write_rows.append((m, {"dmgType": "NON", "currDeg": "a", "prevDeg": "", "detail": "", "pattern": "", "bunrui": "", "bunruiText": ""}, []))
            else:
                write_rows.append((m, None, []))

    total_rows = len(write_rows)

    # 書き込み
    for i, (m, dmg, patterns) in enumerate(write_rows):
        r = start_row + i
        safe_write(ws, r, 2,  m.get("koshu", ""))
        safe_write(ws, r, 5,  m.get("zairyo", ""))
        safe_write(ws, r, 8,  m.get("name", ""))
        safe_write(ws, r, 18, m.get("symbol", ""))
        safe_write(ws, r, 22, m.get("element_no", ""))
        if dmg:
            from openpyxl.styles import Alignment
            safe_write(ws, r, 27, dmg.get("currDeg", ""))
            detail = dmg.get("detail", "")
            unit = extract_unit(detail)
            safe_write(ws, r, 33, remove_units(detail))
            if unit:
                safe_write(ws, r, 44, unit)
            pattern_cell = ws.cell(r, 48)
            if type(pattern_cell).__name__ != "MergedCell":
                pattern_val = ",".join(patterns) if patterns else dmg.get("pattern", "")
                pattern_cell.value = pattern_val
                pattern_cell.alignment = Alignment(horizontal="center", vertical="center")
            dmg_type_cell = ws.cell(r, 53)
            if type(dmg_type_cell).__name__ != "MergedCell":
                dmg_type_cell.value = clean_dmg_type(dmg.get("dmgType", ""))
                dmg_type_cell.alignment = Alignment(horizontal="left", vertical="center")
            bunrui = dmg.get("bunrui", "")
            bunrui_text = dmg.get("bunruiText", "")
            bunrui_val = ("分類" + bunrui + (":" + bunrui_text if bunrui_text else "")) if bunrui else ""
            bunrui_cell = ws.cell(r, 65)
            if type(bunrui_cell).__name__ != "MergedCell":
                bunrui_cell.value = bunrui_val
                bunrui_cell.alignment = Alignment(shrink_to_fit=True)

    return len(write_rows)  # 実際の書き込み行数を返す

def set_print_layout(ws, total_rows: int, start_row: int = 12, rows_per_page: int = 21):
    """
    データ行数に応じて改ページと印刷範囲を設定する。
    - 21行の倍数になるよう印刷範囲を設定（行の追加・削除は行わない）
    - 改ページを21行ごとに挿入
    """
    from openpyxl.worksheet.pagebreak import RowBreak, Break

    # 21行の倍数に切り上げ（最低1ページ分）
    pages = max(1, -(-total_rows // rows_per_page))  # ceil除算
    padded_rows = pages * rows_per_page
    last_data_row = start_row + padded_rows - 1

    # 印刷範囲を更新
    ws.print_area = f"$A$1:$BR${last_data_row}"

    # 改ページを新規作成して設定
    rb = RowBreak()
    page_end = start_row + rows_per_page - 1
    while page_end < last_data_row:
        rb.append(Break(id=page_end, man=True))
        page_end += rows_per_page

    ws.row_breaks = rb

    # 改ページプレビューをデフォルトビューに設定
    ws.sheet_view.view = "pageBreakPreview"


def write_rows_count(members: list, damages: list) -> list:
    """write_membersと同じロジックで書き込み行数を返す（行数計算用）"""
    expanded = expand_members(members)
    dmg_map = {}
    for d in damages:
        if d.get("isNote"):
            continue
        elem_current = d.get("elementNoCurrent", "")
        elem_keys = [e.strip() for e in elem_current.split(",") if e.strip()] if "," in (elem_current or "") else [elem_current]
        for ek in elem_keys:
            key = d.get("symbol", "") + "|" + ek
            dmg_map.setdefault(key, []).append(d)
        # subDamagesの別部材も独立キーで登録（write_membersと同一カウントにする）
        for sub in d.get("subDamages", []):
            sub_sym = sub.get("symbol", "")
            sub_elem = sub.get("elementNoCurrent", "")
            if not sub_sym or sub_sym == d.get("symbol", ""):
                continue
            sub_key = sub_sym + "|" + sub_elem
            sub_entry = {
                "symbol": sub_sym,
                "elementNoCurrent": sub_elem,
                "memberName": sub.get("memberName", ""),
                "dmgType": sub.get("dmgType", ""),
                "currDeg": sub.get("currDeg", ""),
                "prevDeg": sub.get("prevDeg", ""),
                "detail": sub.get("detail", ""),
                "pattern": sub.get("pattern", ""),
                "bunrui": sub.get("bunrui", ""),
                "bunruiText": sub.get("bunruiText", ""),
                "subDamages": [],
                "isNote": False,
            }
            dmg_map.setdefault(sub_key, []).append(sub_entry)
    rows = []
    for m in expanded:
        en = m.get("element_no", "")
        symbol = m.get("symbol", "")
        zairyo = m.get("zairyo", "")
        dmg_list = dmg_map.get(symbol + "|" + en, [])
        targets = get_target_damages(symbol, zairyo)
        if dmg_list:
            # subDamagesも含めたdxf_mapを構築（丸数字キーでユニーク化）
            dxf_keys = set()
            for d in dmg_list:
                k = circled_char(d.get("dmgType", ""))
                if k: dxf_keys.add(k)
                for sub in d.get("subDamages", []):
                    sk = circled_char(sub.get("dmgType", ""))
                    if sk: dxf_keys.add(sk)
            if targets:
                target_keys = {circled_char(t) for t in targets}
                count = len(targets)
                extra = len(dxf_keys - target_keys)
                rows.extend([None] * (count + extra))
            else:
                # targetsなし：実書込と同じく丸数字ユニーク数で計上
                rows.extend([None] * len(dxf_keys))
        else:
            rows.append(None)
    return rows


def extract_unit(quant: str) -> str:
    """定量値から単位を判定して返す"""
    if not quant:
        return ""
    if re.search(r'mm|ｍｍ', quant):
        return "mm"
    if re.search(r'(?<!\d)m(?!m)', quant):
        return "m"
    for unit in ['本', '箇所', '基', '個', '枚', '脚']:
        if unit in quant:
            return unit
    return ""


_UP_S = ['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑩補修・補強材の損傷','⑬遊間の異常','⑱定着部の異常','⑳漏水・滞水','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損']
_UP_C = ['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑨抜け落ち','⑩補修・補強材の損傷','⑪床版ひびわれ','⑫うき','⑬遊間の異常','⑱定着部の異常','⑲変色・劣化','⑳漏水・滞水','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損']
_PB_S = ['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑩補修・補強材の損傷','⑳漏水・滞水','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損']
_PB_C = ['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑩補修・補強材の損傷','⑫うき','⑱定着部の異常','⑲変色・劣化','⑳漏水・滞水','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損']

INSPECTION_TARGETS = {
    'Mg':{'S':_UP_S,'C':_UP_C,'X':[]}, 'Cr':{'S':_UP_S,'C':_UP_C,'X':[]},
    'St':{'S':_UP_S,'C':_UP_C,'X':[]}, 'Ds':{'S':_UP_S,'C':_UP_C,'X':[]},
    'Gb':{'S':_UP_S,'C':_UP_C,'X':[]},
    'Cf':{'S':_UP_S,'C':[],'X':[]}, 'Lu':{'S':_UP_S,'C':[],'X':[]},
    'Ll':{'S':_UP_S,'C':[],'X':[]}, 'Bt':{'S':_UP_S,'C':[],'X':[]},
    'Dt':{'S':_UP_S,'C':[],'X':[]}, 'Pt':{'S':_UP_S,'C':[],'X':[]},
    'Pp':{'S':_UP_S,'C':[],'X':[]}, 'Em':{'S':_UP_S,'C':[],'X':[]},
    'Ar':{'S':[],'C':_UP_C,'X':[]}, 'Sa':{'S':[],'C':_UP_C,'X':[]},
    'Ha':{'S':[],'C':_UP_C,'X':[]}, 'Ca':{'S':[],'C':_UP_C,'X':[]},
    'Rg':{'S':[],'C':_UP_C,'X':[]}, 'Rp':{'S':[],'C':_UP_C,'X':[]},
    'Cn':{'S':['①腐食','⑤防食機能の劣化','㉓変形・欠損'],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑫うき','⑱定着部の異常','⑲変色・劣化','㉓変形・欠損'],'X':[]},
    'Pw':{'S':_PB_S,'C':_PB_C,'X':[]}, 'Pb':{'S':_PB_S,'C':_PB_C,'X':[]},
    'Pc':{'S':_PB_S,'C':_PB_C,'X':[]},
    'Ap':{'S':[],'C':_PB_C,'X':[]}, 'Ac':{'S':[],'C':_PB_C,'X':[]},
    'Aw':{'S':[],'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑩補修・補強材の損傷','⑫うき','⑲変色・劣化','⑳漏水・滞水','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損'],'X':[]},
    'Ff':{'S':['①腐食','②亀裂','⑤防食機能の劣化','㉕沈下・移動・傾斜','㉖洗掘'],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','㉕沈下・移動・傾斜','㉖洗掘'],'X':[]},
    'Bh':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑬遊間の異常','⑯支承部の機能障害','⑳漏水・滞水','㉑異常な音・振動','㉓変形・欠損','㉔土砂詰まり','㉕沈下・移動・傾斜'],
          'C':[],'X':['④破断','⑬遊間の異常','⑯支承部の機能障害','⑲変色・劣化','⑳漏水・滞水','㉑異常な音・振動','㉓変形・欠損','㉔土砂詰まり']},
    'Ba':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑯支承部の機能障害','㉓変形・欠損'],'C':[],'X':[]},
    'Bm':{'S':[],'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑫うき','⑯支承部の機能障害','⑳漏水・滞水','㉓変形・欠損'],'X':[]},
    'Bc':{'S':[],'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑫うき','⑯支承部の機能障害','⑳漏水・滞水','㉓変形・欠損'],'X':[]},
    'Ss':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑬遊間の異常','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損','㉔土砂詰まり'],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑫うき','⑬遊間の異常','⑲変色・劣化','㉓変形・欠損','㉔土砂詰まり'],
          'X':['④破断','⑬遊間の異常','⑲変色・劣化','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損','㉔土砂詰まり']},
    'Sf':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑬遊間の異常','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損','㉔土砂詰まり'],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑫うき','⑬遊間の異常','⑲変色・劣化','㉓変形・欠損','㉔土砂詰まり'],
          'X':['④破断','⑬遊間の異常','⑲変色・劣化','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損','㉔土砂詰まり']},
    'Sd':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑬遊間の異常','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損','㉔土砂詰まり'],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑫うき','⑬遊間の異常','⑲変色・劣化','㉓変形・欠損','㉔土砂詰まり'],
          'X':[]},
    'Ra':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑩補修・補強材の損傷','㉓変形・欠損'],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑩補修・補強材の損傷','⑫うき','⑲変色・劣化','㉓変形・欠損'],'X':[]},
    'Gf':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑩補修・補強材の損傷','㉓変形・欠損'],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑩補修・補強材の損傷','⑫うき','⑲変色・劣化','㉓変形・欠損'],'X':[]},
    'Fg':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑩補修・補強材の損傷','㉓変形・欠損'],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑩補修・補強材の損傷','⑫うき','⑲変色・劣化','㉓変形・欠損'],'X':[]},
    'Ej':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','⑬遊間の異常','⑭路面の凹凸','⑳漏水・滞水','㉑異常な音・振動','㉓変形・欠損','㉔土砂詰まり'],
          'C':['⑥ひびわれ','⑫うき','㉑異常な音・振動','㉓変形・欠損'],
          'X':['⑬遊間の異常','⑭路面の凹凸','⑲変色・劣化','⑳漏水・滞水','㉑異常な音・振動','㉓変形・欠損','㉔土砂詰まり']},
    'Dr':{'S':['①腐食','④破断','⑤防食機能の劣化','⑲変色・劣化','⑳漏水・滞水','㉓変形・欠損','㉔土砂詰まり'],
          'C':[],'X':['④破断','⑲変色・劣化','⑳漏水・滞水','㉓変形・欠損','㉔土砂詰まり']},
    'Dp':{'S':['①腐食','④破断','⑤防食機能の劣化','⑲変色・劣化','⑳漏水・滞水','㉓変形・欠損','㉔土砂詰まり'],
          'C':[],'X':['④破断','⑲変色・劣化','⑳漏水・滞水','㉓変形・欠損','㉔土砂詰まり']},
    'Ip':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損'],'C':[],'X':[]},
    'Ut':{'S':['①腐食','②亀裂','③ゆるみ・脱落','④破断','⑤防食機能の劣化','㉑異常な音・振動','㉒異常なたわみ','㉓変形・欠損'],'C':[],'X':[]},
    'Cu':{'S':[],
          'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑫うき','⑲変色・劣化','㉓変形・欠損'],'X':[]},
    'Pm':{'S':[],
          'C':['⑭路面の凹凸','⑮舗装の異常','㉔土砂詰まり'],
          'X':['⑭路面の凹凸','⑮舗装の異常','㉔土砂詰まり']},
    'Ww':{'S':[],'C':['⑥ひびわれ','⑦剥離・鉄筋露出','⑧漏水・遊離石灰','⑲変色・劣化','㉓変形・欠損','㉕沈下・移動・傾斜'],'X':[]},
}


def circled_char(s: str) -> str:
    """文字列の先頭の丸数字を返す"""
    if not s: return ""
    m = re.match(r'^([①-⑳㉑-㉛])', s)
    return m.group(1) if m else ""


def get_target_damages(symbol: str, zairyo: str) -> list:
    """部材記号と材料から対象損傷種類リストを返す"""
    entry = INSPECTION_TARGETS.get(symbol)
    if not entry:
        return []
    if zairyo == 'S':
        key = 'S'
    elif zairyo == 'C':
        key = 'C'
    elif zairyo:
        key = 'X'
    else:
        return []
    return entry.get(key, [])


def remove_units(quant: str) -> str:
    """定量値から単位（mm/m/本等）を除去する"""
    if not quant:
        return quant
    result = re.sub(r'mm|ｍｍ', '', quant)
    result = re.sub(r'(\d)m(?!\d)', r'\1', result)
    for unit in ['本', '箇所', '基', '個', '枚', '脚']:
        result = result.replace(unit, '')
    return result.strip()


def clean_dmg_type(dmg_type: str) -> str:
    """損傷の種類から先頭の丸数字を除去する"""
    if not dmg_type:
        return ""
    return re.sub(r'^[①-⑳㉑-㉛]', '', dmg_type).strip()


# -------------------------------------------------------
# その３-４：損傷程度の評価結果総括
# -------------------------------------------------------

# 下部工（橋台・橋脚・基礎）は要素番号の後2桁が部材番号
LOWER_STRUCTURE_SYMBOLS = {'Ap', 'Ac', 'Aw', 'Pw', 'Pb', 'Pc', 'Ff'}
MAJOR_SYMBOLS = {'Mg', 'Cr', 'St', 'Ds', 'Gb', 'Cf', 'Lu', 'Ll', 'Bt', 'Dt', 'Pt', 'Pp', 'Em',
                 'Ar', 'Sa', 'Ha', 'Ca', 'Rg', 'Rp', 'Cn', 'Pw', 'Pb', 'Pc', 'Ap', 'Ac', 'Aw', 'Ff',
                 'Bh', 'Ba', 'Bm', 'Bc'}
# 損傷程度の序列（最大値を採用するため）
_DEG_RANK_34 = {'a': 1, 'b': 2, 'c': 3, 'd': 4, 'e': 5}


def buzai_no_34(symbol: str, element_no: str) -> str:
    sym = symbol
    en = (element_no or '').strip()
    if sym not in MAJOR_SYMBOLS:
        return '00'
    if sym == 'Ds':
        return '00'
    if len(en) == 4 and en.isdigit():
        return en[2:] if sym in LOWER_STRUCTURE_SYMBOLS else en[:2]
    if len(en) == 2 and en.isdigit():
        return en
    return '00'


def build_summary_rows_34(members: list, damages: list, prev_items: list = None) -> list:
    """部材番号単位で損傷を集約し、書き込み行リストを返す。
    返り値: [(member_or_None, 部材番号, 今回[(名,程度)...], 前回[(名,程度)...], NGフラグ), ...]
    - 同一損傷種類は最大評価のみ
    - 並びは損傷番号①→㉖の順
    - 損傷なしで対象損傷がある部材は NON(a)
    - 4件を超える場合は次行に継続（部材欄は空欄）
    """
    circle_order = {ch: i + 1 for i, ch in enumerate(CIRCLE_NUMS)}
    expanded = expand_members(members)

    # 損傷データを symbol|要素番号 でマッピング（その３-３と同一キー）
    dmg_map = {}
    for d in damages:
        if d.get("isNote"):
            continue
        elem_current = d.get("elementNoCurrent", "")
        elem_keys = [e.strip() for e in elem_current.split(",") if e.strip()] if "," in (elem_current or "") else [elem_current]
        for ek in elem_keys:
            key = d.get("symbol", "") + "|" + ek
            dmg_map.setdefault(key, []).append(d)

    groups = {}
    order = []
    for m in expanded:
        symbol = m.get("symbol", "")
        en = m.get("element_no", "")
        symbol_n = symbol
        zairyo = m.get("zairyo", "")
        bn = buzai_no_34(symbol_n, en)
        gkey = symbol_n + "|" + bn + "|" + zairyo
        if gkey not in groups:
            groups[gkey] = {"m": m, "bn": bn, "dmg": {}}
            order.append(gkey)
        g = groups[gkey]
        dmg_list = dmg_map.get(symbol + "|" + en, [])
        for d in dmg_list:
            entries = [d] + list(d.get("subDamages", []))
            for ent in entries:
                ck = circled_char(ent.get("dmgType", ""))
                if not ck:
                    continue
                deg = (ent.get("currDeg") or "").strip()
                cur = g["dmg"].get(ck)
                if cur is None or _DEG_RANK_34.get(deg, 0) > _DEG_RANK_34.get(cur[0], 0):
                    g["dmg"][ck] = (deg, clean_dmg_type(ent.get("dmgType", "")))

    # 前回データを 記号|部材番号 でマッピング
    prev_map = {}
    prev_order = []
    for it in (prev_items or []):
        sym_p = it.get("symbol", "")
        pk = sym_p + "|" + it.get("buzai_no", "") + "|" + it.get("zairyo", "")
        if pk not in prev_map:
            prev_map[pk] = it
            prev_order.append(pk)

    rows = []
    matched_prev = set()
    for gkey in order:
        g = groups[gkey]
        m = g["m"]
        items = sorted(g["dmg"].items(), key=lambda kv: circle_order.get(kv[0], 99))
        pairs = [(name, deg) for _, (deg, name) in items]
        curr_names = {_norm_dmg_name_34(name) for name, _ in pairs}
        if not pairs:
            targets = get_target_damages(m.get("symbol", ""), m.get("zairyo", ""))
            if targets:
                pairs = [("NON", "a")]

        # 前回損傷（そのまま転記・修正禁止）
        prev_it = prev_map.get(gkey)
        if prev_it is not None:
            matched_prev.add(gkey)
            prev_pairs = [(n, d) for n, d in prev_it.get("damages", [])]
        else:
            prev_pairs = [("NON", "a")]

        # NG判定：前回損傷（NON以外）が今回データに存在しない場合
        ng = ""
        for n, _d in prev_pairs:
            if n == "NON":
                continue
            if _norm_dmg_name_34(n) not in curr_names:
                ng = "NG"
                break

        n_rows = max(1, -(-len(pairs) // 4), -(-len(prev_pairs) // 4))
        for i in range(n_rows):
            rows.append((
                m if i == 0 else None,
                g["bn"],
                pairs[i * 4:(i + 1) * 4],
                prev_pairs[i * 4:(i + 1) * 4],
                ng if i == 0 else "",
            ))

    return rows


def _norm_dmg_name_34(s: str) -> str:
    """損傷名の表記ゆれを正規化（中黒・空白）"""
    return (s or "").replace("\uff65", "\u30fb").replace("･", "・").replace(" ", "").replace("\u3000", "").strip()


def _parse_prev_dmg_text_34(text: str) -> list:
    """「名称（程度）,名称（程度）」形式を [(名称, 程度), ...] に分解"""
    pairs = []
    if not text:
        return pairs
    for part in re.split(r'[,，]', str(text)):
        part = part.strip()
        if not part:
            continue
        mt = re.match(r'^(.+?)[（(]([a-eA-E])[）)]\s*$', part)
        if mt:
            pairs.append((mt.group(1).strip(), mt.group(2).lower()))
        else:
            pairs.append((part, ""))
    return pairs


def extract_prev_summary_34(xls_path: str) -> dict:
    """前回調書から損傷程度の評価結果総括（その13等）を抽出する。
    シートはB2セルの文言で特定（シート名に非依存）、径間はZ2セル。
    返り値: {径間番号(str): {"date": 点検日, "items": [ {koshu,zairyo,name,symbol,buzai_no,damages:[[名称,程度],...]} ]}}
    """
    wb = open_workbook_any(xls_path)
    result = {}
    for sname in wb.sheet_names():
        sh = wb.sheet_by_name(sname)
        if sh.nrows < 10 or sh.ncols < 22:
            continue
        b2 = str(sh.cell_value(1, 1)).strip()
        if "損傷程度の評価結果総括" not in b2:
            continue
        z2 = sh.cell_value(1, 25)
        try:
            span_no = str(int(float(z2)))
        except (ValueError, TypeError):
            continue
        # 点検日（前回調書の「今回定期点検」点検日 ＝ こちらの前回点検日）AI8
        date_val = str(sh.cell_value(7, 34)).strip() if sh.ncols > 34 else ""
        entry = result.setdefault(span_no, {"date": date_val, "items": []})
        if date_val and not entry.get("date"):
            entry["date"] = date_val
        for r in range(9, sh.nrows):
            symbol = str(sh.cell_value(r, 13)).strip()
            if not symbol:
                continue
            bn_raw = sh.cell_value(r, 16)
            if isinstance(bn_raw, float):
                bn = str(int(bn_raw)).zfill(2)
            else:
                bn = str(bn_raw).strip()
            dmg_text = str(sh.cell_value(r, 21)).strip()
            entry["items"].append({
                "koshu":  str(sh.cell_value(r, 1)).strip(),
                "zairyo": str(sh.cell_value(r, 4)).strip(),
                "name":   str(sh.cell_value(r, 7)).strip(),
                "symbol": symbol,
                "buzai_no": bn,
                "damages": [[n, d] for n, d in _parse_prev_dmg_text_34(dmg_text)],
            })
    return result


def write_summary_34(ws, rows: list, start_row: int = 12):
    """その３-４のデータ行を書き込む（1ページ＝最大18行：行12〜29）"""
    NAME_COLS = [22, 28, 34, 40]       # 今回：V, AB, AH, AN
    DEG_COLS  = [27, 33, 39, 45]       # 今回：AA, AG, AM, AS
    PREV_NAME_COLS = [46, 52, 58, 64]  # 前回：AT, AZ, BF, BL
    PREV_DEG_COLS  = [51, 57, 63, 69]  # 前回：AY, BE, BK, BQ
    NG_COL = 75                        # BW
    for i, (m, bn, dmgs, prevs, ng) in enumerate(rows):
        r = start_row + i
        if m is not None:
            safe_write(ws, r, 2,  m.get("koshu", ""))
            safe_write(ws, r, 5,  m.get("zairyo", ""))
            safe_write(ws, r, 8,  m.get("name", ""))
            safe_write(ws, r, 15, m.get("symbol", ""))
            safe_write(ws, r, 18, bn)
        for j, (nm, deg) in enumerate(dmgs[:4]):
            safe_write(ws, r, NAME_COLS[j], nm)
            safe_write(ws, r, DEG_COLS[j], deg)
        for j, (nm, deg) in enumerate(prevs[:4]):
            safe_write(ws, r, PREV_NAME_COLS[j], nm)
            safe_write(ws, r, PREV_DEG_COLS[j], deg)
        if ng:
            safe_write(ws, r, NG_COL, ng)


def write_template(data: dict, output_path: str):
    shutil.copy(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path, keep_vba=False)

    # --- その１：ヘッダーのみ ---
    ws1 = wb["その１"]
    ws1["H5"]  = data["フリガナ"]
    ws1["H6"]  = data["橋梁名"]
    ws1["AE5"] = data["路線名"]
    ws1["AQ5"] = data["管理者"]

    # 損傷データ（径間・要素番号で部材行と紐付け）
    damages_data = data.get("damages_data", [])

    # --- その３-３：径間別に部材リストを書き込む ---
    spans = data.get("members_by_span", [])
    base_sheet_name = "その３-３"

    # クリーンテンプレを事前に保持（書き込み前）
    ws33_tmpl = wb[base_sheet_name]
    _ws33_clean = wb.copy_worksheet(ws33_tmpl)
    _ws33_clean.title = "__clean_33__"
    _ws33_clean.sheet_state = "hidden"
    base_33_insert_pos = wb.sheetnames.index(base_sheet_name)

    for idx, span in enumerate(spans):
        span_no = span.get("span_no", idx + 1)
        members = span.get("major", []) + span.get("other", [])
        span_damages = [d for d in damages_data if d.get("spanNo") == span_no]

        # 必要ページ数を計算
        needed = len(write_rows_count(members, span_damages))
        pages = max(1, -(-needed // 21))
        total_needed = pages * 21

        # クリーンテンプレから空欄シートを生成
        sheet_label = base_sheet_name if idx == 0 else f"{base_sheet_name} ({idx + 1})"
        if sheet_label in wb.sheetnames:
            del wb[sheet_label]
        ws = wb.copy_worksheet(_ws33_clean)
        ws.title = sheet_label
        target_idx = base_33_insert_pos + idx
        current_idx = wb.sheetnames.index(sheet_label)
        wb.move_sheet(sheet_label, offset=target_idx - current_idx)
        ws.sheet_view.showGridLines = False
        ws.print_title_rows = "$1:$11"

        # 必要ページ分の行をクリーンテンプレ12行目からコピーして準備
        for i in range(1, total_needed):
            dst_row = 12 + i
            copy_row_format(ws, 12, dst_row)

        # 必要な情報を書き込む
        write_sheet_header(ws, data, span_no)
        actual_rows = write_members(ws, members, span_damages, span_no)
        set_print_layout(ws, actual_rows)

    if "__clean_33__" in wb.sheetnames:
        del wb["__clean_33__"]

    # --- その３-４：損傷程度の評価結果総括（部材番号単位で集約） ---
    base34 = "その３-４"
    prev_summary = data.get("prev_summary", {}) or {}
    ROWS_PER_PAGE_34 = 20  # 行12〜31が1ページ分（印刷タイトル$1:$11）
    DATA_COLS_34 = [2, 5, 8, 15, 18, 22, 27, 28, 33, 34, 39, 40, 45,
                    46, 51, 52, 57, 58, 63, 64, 69, 75]
    for idx, span in enumerate(spans):
        span_no = span.get("span_no", idx + 1)
        members = span.get("major", []) + span.get("other", [])
        span_damages = [d for d in damages_data if d.get("spanNo") == span_no]
        prev34 = prev_summary.get(str(span_no), {}) if isinstance(prev_summary, dict) else {}
        rows34 = build_summary_rows_34(members, span_damages, prev34.get("items", []))

        if idx == 0:
            ws34 = wb[base34]
        else:
            sheet_name = f"{base34} ({idx + 1})"
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws34 = wb.copy_worksheet(wb[base34])
            ws34.title = sheet_name
            base_idx = wb.sheetnames.index(base34)
            current_idx = wb.sheetnames.index(sheet_name)
            wb.move_sheet(sheet_name, offset=(base_idx + idx) - current_idx)
            ws34.sheet_view.showGridLines = False
            ws34.print_title_rows = "$1:$11"
            # コピー元のデータ行をクリア
            for r in range(12, ws34.max_row + 1):
                for c in DATA_COLS_34:
                    safe_write(ws34, r, c, None)

        write_sheet_header(ws34, data, span_no)
        prev_date = prev34.get("date", "")
        if prev_date:
            safe_write(ws34, 10, 63, prev_date)  # BK10：前回点検日

        # 必要行数からページ数（20行の倍数）を確保し、不足行はテンプレ行をコピー
        needed = len(rows34)
        pages = max(1, -(-needed // ROWS_PER_PAGE_34))
        total_needed = pages * ROWS_PER_PAGE_34
        for i in range(1, total_needed):
            dst_row = 12 + i
            will_have_data = i < needed  # この行に実データが書き込まれる
            if ws34.max_row < dst_row or will_have_data or (ws34.cell(dst_row, 2).value is None and ws34.cell(dst_row, 8).value is None):
                copy_row_format(ws34, 12, dst_row)

        write_summary_34(ws34, rows34)
        # 改ページ・印刷範囲・改ページプレビューを設定
        set_print_layout(ws34, needed, start_row=12, rows_per_page=ROWS_PER_PAGE_34)

    # --- その１：現地状況写真を書き込む ---
    situation_photos = data.get("situation_photos", [])
    photos_b64_keys = list(data.get("photos_base64", {}).keys())
    print(f"[DEBUG] photos_base64 count={len(photos_b64_keys)} keys={photos_b64_keys[:3]}")
    print(f"[DEBUG] situation_photos assignedKeys={[p.get('assignedKey') for p in situation_photos[:3]]}")
    if situation_photos:
        PAGE_ROWS_1 = 8  # 行10〜17が1ページ分
        COL_VALS_1 = [
            {"写真番号": "H", "径間番号": "R", "メモ": "H"},
            {"写真番号": "AD", "径間番号": "AN", "メモ": "AD"},
            {"写真番号": "AZ", "径間番号": "BJ", "メモ": "AZ"},
        ]
        VALUE_COLS_1 = ['H','AD','AZ','R','AN','BJ']
        from openpyxl.utils import column_index_from_string as c2i_1
        from openpyxl.utils import get_column_letter as gcl_1
        from openpyxl.worksheet.pagebreak import Break as Break1
        from copy import copy as _copy1

        def copy_page_1(ws, dst_start):
            """行10〜17（1ページ分）をdst_startにコピー"""
            src_s, src_e = 10, 17
            num = src_e - src_s + 1
            # dst範囲の結合を解除
            for m in list(ws.merged_cells.ranges):
                if dst_start <= m.min_row <= dst_start + num - 1:
                    ws.unmerge_cells(str(m))
            # 書式・値コピー
            for r in range(num):
                sr, dr = src_s + r, dst_start + r
                for c in range(1, ws.max_column + 1):
                    src = ws.cell(sr, c)
                    dst = ws.cell(dr, c)
                    if type(dst).__name__ == "MergedCell":
                        continue
                    if src.has_style:
                        dst.font = _copy1(src.font)
                        dst.border = _copy1(src.border)
                        dst.fill = _copy1(src.fill)
                        dst.number_format = src.number_format
                        dst.alignment = _copy1(src.alignment)
                    dst.value = src.value
                ws.row_dimensions[dr].height = ws.row_dimensions[sr].height
            # 結合コピー
            offset = dst_start - src_s
            for m in list(ws.merged_cells.ranges):
                if src_s <= m.min_row and m.max_row <= src_e:
                    nr = f"{gcl_1(m.min_col)}{m.min_row+offset}:{gcl_1(m.max_col)}{m.max_row+offset}"
                    try:
                        ws.merge_cells(nr)
                    except Exception:
                        pass
            # コピー先の値セルをクリア
            for r in range(num):
                row = dst_start + r
                for col_letter in VALUE_COLS_1:
                    c = ws.cell(row, c2i_1(col_letter))
                    if type(c).__name__ != 'MergedCell':
                        c.value = None

        base_sheet_1 = "その１"
        span_nos_1 = sorted(set(p.get("spanNo") for p in situation_photos if p.get("spanNo") is not None))
        if not span_nos_1:
            span_nos_1 = [1]
        sheet_1_insert_pos = wb.sheetnames.index(base_sheet_1)

        # STEP1: 径間数分シートをコピー
        ws1_sheets = {}
        for s_idx, span_no in enumerate(span_nos_1):
            if s_idx == 0:
                ws1_sheets[span_no] = wb[base_sheet_1]
            else:
                sheet_label = f"{base_sheet_1} ({s_idx + 1})"
                if sheet_label in wb.sheetnames:
                    del wb[sheet_label]
                ws1 = wb.copy_worksheet(wb[base_sheet_1])
                ws1.title = sheet_label
                target_idx = sheet_1_insert_pos + s_idx
                current_idx = wb.sheetnames.index(sheet_label)
                wb.move_sheet(sheet_label, offset=target_idx - current_idx)
                ws1.sheet_view.showGridLines = False
                ws1.print_title_rows = "$1:$9"
                ws1_sheets[span_no] = ws1

        # STEP2: 各シートで必要ページ分コピー追加・値書き込み
        for s_idx, span_no in enumerate(span_nos_1):
            ws1 = ws1_sheets[span_no]
            span_photos = [p for p in situation_photos if p.get("spanNo") == span_no]
            pages1 = max(1, -(-len(span_photos) // 6))

            write_sheet_header(ws1, data, span_no, write_span_no=False)
            ws1.sheet_view.view = "pageBreakPreview"

            # 1ページ目の値セルをクリア
            for row in range(10, 18):
                for col_letter in VALUE_COLS_1:
                    c = ws1.cell(row, c2i_1(col_letter))
                    if type(c).__name__ != 'MergedCell':
                        c.value = None

            # 2ページ目以降を追加
            for page in range(1, pages1):
                copy_page_1(ws1, 10 + PAGE_ROWS_1 * page)

            # 改ページ・印刷範囲
            for page in range(1, pages1):
                ws1.row_breaks.append(Break1(id=10 + PAGE_ROWS_1 * page - 1))
            ws1.print_area = f"A1:{gcl_1(ws1.max_column)}{9 + PAGE_ROWS_1 * pages1}"

            # 値を書き込む
            photos_base64 = data.get("photos_base64", {})
            for i, photo in enumerate(span_photos):
                page = i // 6
                pos = i % 6
                col = COL_VALS_1[pos % 3]
                base = 10 + PAGE_ROWS_1 * page + (pos // 3) * 4
                ws1[f"{col['写真番号']}{base}"] = photo.get("photoNum", "")
                ws1[f"{col['径間番号']}{base}"] = photo.get("spanNo", "")
                ws1[f"{col['メモ']}{base + 1}"] = photo.get("memo", "")

                # 写真貼り付け
                assigned_key = photo.get("assignedKey", "")
                if assigned_key and assigned_key in photos_base64:
                    try:
                        import base64 as _b64, io as _io
                        from PIL import Image as _PIL
                        from openpyxl.drawing.image import Image as _XLImg
                        b64_data = photos_base64[assigned_key].split(",")[-1]
                        img_bytes = _b64.b64decode(b64_data)
                        pil_img = _PIL.open(_io.BytesIO(img_bytes))
                        target_h = 220
                        target_w = int(target_h * 4 / 3)
                        pil_img = pil_img.resize((target_w, target_h), _PIL.LANCZOS)
                        img_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                        pil_img.save(img_tmp.name, "PNG")
                        img_tmp.close()
                        xl_img = _XLImg(img_tmp.name)
                        # サイズ固定・中央配置（OneCellAnchor）
                        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                        from openpyxl.utils.units import pixels_to_EMU
                        target_h_px = int(220 * 0.95)
                        target_w_px = int(target_h_px * 4 / 3)
                        pil_img = pil_img.resize((target_w_px, target_h_px), _PIL.LANCZOS)
                        pil_img.save(img_tmp.name, "PNG")
                        xl_img = _XLImg(img_tmp.name)
                        row_offset_px = int((220 - target_h_px) / 2)
                        # D〜Y中央: K列中心に写真を配置
                        # I列(0-indexed=8)から53pxオフセット
                        anchor_cols_idx = [5, 27, 49]  # F=5, AB=27, AX=49 (0-indexed)
                        col_idx = anchor_cols_idx[pos % 3]
                        row_idx = base + 2
                        marker = AnchorMarker(
                            col=col_idx, colOff=pixels_to_EMU(int((2112 - target_w_px) / 2)),
                            row=row_idx, rowOff=pixels_to_EMU(row_offset_px)
                        )
                        xl_img.anchor = OneCellAnchor(_from=marker,
                            ext=openpyxl.drawing.xdr.XDRPositiveSize2D(
                                pixels_to_EMU(target_w_px), pixels_to_EMU(target_h_px)
                            )
                        )
                        ws1.add_image(xl_img)
                    except Exception as e:
                        print(f"写真貼り付けエラー: {e}")
    # --- その３-２：損傷写真シートを書き込む ---
    base_sheet_32 = "その３-２"
    span_nos = sorted(set(d.get("spanNo") for d in damages_data if d.get("spanNo") is not None))
    if not span_nos:
        span_nos = [1]

    # 左列・中列・右列の値セル列
    COL_VALS = [
        {"写真番号": "H", "部材名": "H", "要素番号": "S", "損傷の種類": "H", "損傷程度": "S", "前回損傷程度": "U", "メモ": "U"},
        {"写真番号": "AD", "部材名": "AD", "要素番号": "AO", "損傷の種類": "AD", "損傷程度": "AO", "前回損傷程度": "AQ", "メモ": "AQ"},
        {"写真番号": "AZ", "部材名": "AZ", "要素番号": "BK", "損傷の種類": "AZ", "損傷程度": "BK", "前回損傷程度": "BM", "メモ": "BM"},
    ]

    def write_damage_entry(ws32, dmg, col, row_offset):
        """1件の損傷データを指定列・行オフセットに書き込む"""
        base = row_offset
        raw_label = dmg.get("photoLabel", "")
        import re as _re
        photo_num = _re.sub(r'^[^0-9]*', '', str(raw_label))
        ws32[f"{col['写真番号']}{base}"] = int(photo_num) if photo_num.isdigit() else raw_label
        ws32[f"{col['部材名']}{base + 1}"] = dmg.get("memberName", "")
        ws32[f"{col['要素番号']}{base + 1}"] = dmg.get("elementNoCurrent", "")
        ws32[f"{col['損傷の種類']}{base + 2}"] = dmg.get("dmgType", "")
        ws32[f"{col['損傷程度']}{base + 2}"] = dmg.get("currDeg", "")
        prev_no = dmg.get("damageNoPrev", "") or dmg.get("damage_no_prev", "")
        prev_deg = dmg.get("prevDeg", "")
        prev_val = f"{prev_no}-{prev_deg}" if prev_no and prev_deg else prev_deg
        ws32[f"{col['前回損傷程度']}{base + 4}"] = prev_val
        memo_type = dmg.get("memoType", "normal") or "normal"
        memo_templates = data.get("memo_templates") or {}
        if memo_type in ("repaired", "none") and memo_templates.get(memo_type):
            member_name = dmg.get("memberName", "")
            dmg_type = dmg.get("dmgType", "")
            strip_circle = lambda s: re.sub(r'^[①-⑳㉑-㉛]', '', s or '').strip()
            subs = dmg.get("subDamages", [])
            all_types = [strip_circle(dmg_type)] + [strip_circle(s.get("dmgType","")) for s in subs if s.get("dmgType")]
            memo = memo_templates[memo_type].replace("{部材名}", member_name).replace("{損傷種類}", "、".join(filter(None, all_types)))
        else:
            memo = dmg.get("memo", "") or ""
        subs = dmg.get("subDamages", [])
        if subs and memo_type == "normal":
            sub_lines = []
            for s in subs:
                sym_m = re.match(r'([①-⑳㉑-㉛])', s.get("dmgType","") or "")
                sym = sym_m.group(1) if sym_m else ""
                deg = s.get("currDeg","") or ""
                if sym and deg:
                    sub_lines.append(f"{sym}-{deg}")
            if sub_lines:
                memo = (memo + "\n" if memo else "") + "\n".join(sub_lines)
        ws32[f"{col['メモ']}{base + 6}"] = memo

    # テンプレ1ページ目（書込み前）をメモリに退避
    ws32_tmpl = wb[base_sheet_32]
    # ws32_tmplを別名で保護（元シート削除時の参照破壊を防ぐ）
    _ws32_clean = wb.copy_worksheet(ws32_tmpl)
    _ws32_clean.title = "__clean_32__"
    _ws32_clean.sheet_state = "hidden"
    ws32_tmpl = _ws32_clean
    template_rows_32 = {}
    for _r in range(11, 35):
        template_rows_32[_r] = {}
        for _c in range(1, ws32_tmpl.max_column + 1):
            _cell = ws32_tmpl.cell(_r, _c)
            if type(_cell).__name__ != 'MergedCell':
                template_rows_32[_r][_c] = _cell.value

    def copy_rows_32(ws32, src_start, src_end, dst_start):
        """11〜34行（1ページ分）を指定行にコピー"""
        from copy import copy
        from openpyxl.utils import get_column_letter
        num_rows = src_end - src_start + 1

        # 既存のdst範囲の結合を先にまとめて解除
        for merged in list(ws32.merged_cells.ranges):
            if dst_start <= merged.min_row <= dst_start + num_rows - 1:
                ws32.unmerge_cells(str(merged))

        # 書式・値コピー（ws32_tmplから）
        for r in range(num_rows):
            src_row = src_start + r
            dst_row = dst_start + r
            for col in range(1, ws32_tmpl.max_column + 1):
                src = ws32_tmpl.cell(src_row, col)
                dst = ws32.cell(dst_row, col)
                if type(dst).__name__ == "MergedCell":
                    continue
                if type(src).__name__ == "MergedCell":
                    # アンカーから書式取得
                    anchor = None
                    for m in ws32_tmpl.merged_cells.ranges:
                        if m.min_row <= src_row <= m.max_row and m.min_col <= col <= m.max_col:
                            anchor = ws32_tmpl.cell(m.min_row, m.min_col)
                            break
                    src = anchor or src
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                    dst.number_format = src.number_format
                    dst.alignment = copy(src.alignment)
                dst.value = template_rows_32.get(src_row, {}).get(col, None)
            ws32.row_dimensions[dst_row].height = ws32_tmpl.row_dimensions[src_row].height

        # 結合セルをsrc→dstにコピー（複数行にまたがるものも含む）
        offset = dst_start - src_start
        for merged in list(ws32.merged_cells.ranges):
            if src_start <= merged.min_row and merged.max_row <= src_end:
                new_range = (f"{get_column_letter(merged.min_col)}{merged.min_row + offset}:"
                             f"{get_column_letter(merged.max_col)}{merged.max_row + offset}")
                try:
                    ws32.merge_cells(new_range)
                except Exception:
                    pass

    sheet_32_insert_pos = wb.sheetnames.index(base_sheet_32)
    sheet_32_counter = 0

    for s_idx, span_no in enumerate(span_nos):
        span_dmgs = [d for d in damages_data if d.get("spanNo") == span_no and d.get("photoLabel") and not d.get("isNote")]
        pages = max(1, -(-len(span_dmgs) // 6))

        # シートを作成（径間ごとに1シート、全径間をws32_tmplからコピー）
        sheet_label = base_sheet_32 if s_idx == 0 else f"{base_sheet_32} ({s_idx + 1})"
        if sheet_label in wb.sheetnames:
            del wb[sheet_label]
        ws32 = wb.copy_worksheet(ws32_tmpl)
        ws32.title = sheet_label
        target_idx = sheet_32_insert_pos + s_idx
        current_idx = wb.sheetnames.index(sheet_label)
        wb.move_sheet(sheet_label, offset=target_idx - current_idx)
        ws32.sheet_view.showGridLines = False
        ws32.print_title_rows = "$1:$10"
        ws32["H5"] = None
        ws32["H6"] = None
        ws32["AE5"] = None
        ws32["AQ5"] = None

        sheet_32_counter += 1
        write_sheet_header(ws32, data, span_no)
        # 改ページプレビュー設定
        ws32.sheet_view.view = "pageBreakPreview"

        # ページ数分11〜34行をコピーして追加
        PAGE_ROWS = 24  # 11〜34行 = 24行
        from openpyxl.worksheet.pagebreak import Break
        for page in range(1, pages):
            dst_start = 11 + PAGE_ROWS * page
            copy_rows_32(ws32, 11, 34, dst_start)
            break_row = dst_start - 1
            ws32.row_breaks.append(Break(id=break_row))

        # 印刷範囲を設定（行10+24×pages）
        last_row = 10 + PAGE_ROWS * pages
        from openpyxl.utils import get_column_letter
        max_col = get_column_letter(ws32.max_column)
        ws32.print_area = f"A1:{max_col}{last_row}"

        # 各損傷データを書き込む
        photos_base64_32 = data.get("photos_base64", {})
        for i, dmg in enumerate(span_dmgs):
            page = i // 6
            pos_in_page = i % 6  # 0〜5
            row_in_page = (pos_in_page // 3)  # 0=上段, 1=下段
            col_in_page = pos_in_page % 3     # 0=左, 1=中, 2=右
            base_row = 11 + PAGE_ROWS * page + row_in_page * 12
            col = COL_VALS[col_in_page]
            write_damage_entry(ws32, dmg, col, base_row)

            # 写真貼り付け
            assigned_key = dmg.get("assignedKey", "")
            print(f"[DEBUG 32] i={i} assignedKey={assigned_key!r} in_photos={assigned_key in photos_base64_32}")
            if assigned_key and assigned_key in photos_base64_32:
                try:
                    import base64 as _b64, io as _io
                    from PIL import Image as _PIL
                    from openpyxl.drawing.image import Image as _XLImg
                    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                    from openpyxl.utils.units import pixels_to_EMU
                    b64_data = photos_base64_32[assigned_key].split(",")[-1]
                    img_bytes = _b64.b64decode(b64_data)
                    pil_img = _PIL.open(_io.BytesIO(img_bytes))
                    target_h_px = int(min(229, 2112) * 0.95 * 0.95)
                    target_w_px = int(target_h_px * 4 / 3)
                    pil_img = pil_img.resize((target_w_px, target_h_px), _PIL.LANCZOS)
                    img_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                    pil_img.save(img_tmp.name, "PNG")
                    img_tmp.close()
                    xl_img = _XLImg(img_tmp.name)
                    anchor_cols_idx_32 = [3, 25, 47]  # D=3, Z=25, AV=47 (0始まり)
                    col_idx_32 = anchor_cols_idx_32[col_in_page]
                    # 写真開始行: 上段=row14(idx=13), 下段=row27(idx=26), 0始まり
                    row_idx_32 = 11 + PAGE_ROWS * page + row_in_page * 13 + 3 - 2 + (1 if row_in_page == 0 else 0)
                    marker = AnchorMarker(
                        col=col_idx_32, colOff=pixels_to_EMU(10),
                        row=row_idx_32, rowOff=pixels_to_EMU(15)
                    )
                    xl_img.anchor = OneCellAnchor(_from=marker,
                        ext=openpyxl.drawing.xdr.XDRPositiveSize2D(
                            pixels_to_EMU(target_w_px), pixels_to_EMU(target_h_px)
                        )
                    )
                    ws32.add_image(xl_img)
                except Exception as e:
                    print(f"その３-２写真貼り付けエラー: {e}")

        # === NON写真をこの径間の損傷写真の後に追記 ===
        non_photos = data.get("non_photos", [])
        span_nons = [n for n in non_photos if (n.get("spanNo") or 1) == span_no and n.get("photoLabel")]
        if span_nons:
            # 部材名解決用マップ（記号→名称）
            sym_name = {}
            for sp in data.get("members_by_span", []):
                if (sp.get("span_no") or 1) != span_no:
                    continue
                for m in (sp.get("major", []) + sp.get("other", [])):
                    if m.get("symbol"):
                        sym_name[m["symbol"]] = m.get("name", "")

            # 損傷写真の最終ページの次ページからNON写真を開始
            dmg_pages = max(1, -(-len(span_dmgs) // 6))
            non_start_page = dmg_pages  # 0始まりページ番号で次ページ
            non_total_pages = -(-len(span_nons) // 6)
            total_pages = dmg_pages + non_total_pages

            # 追加ページ分の行をコピー＋改ページ
            for page in range(dmg_pages, total_pages):
                dst_start = 11 + PAGE_ROWS * page
                copy_rows_32(ws32, 11, 34, dst_start)
                ws32.row_breaks.append(Break(id=dst_start - 1))

            # 印刷範囲を全ページ分に更新
            last_row = 10 + PAGE_ROWS * total_pages
            max_col = get_column_letter(ws32.max_column)
            ws32.print_area = f"A1:{max_col}{last_row}"

            for ni, non in enumerate(span_nons):
                page = non_start_page + ni // 6
                pos_in_page = ni % 6
                row_in_page = pos_in_page // 3
                col_in_page = pos_in_page % 3
                base_row = 11 + PAGE_ROWS * page + row_in_page * 12
                col = COL_VALS[col_in_page]

                # 写真番号
                raw_label = non.get("photoLabel", "")
                import re as _re2
                pnum = _re2.sub(r'^[^0-9]*', '', str(raw_label))
                ws32[f"{col['写真番号']}{base_row}"] = int(pnum) if pnum.isdigit() else raw_label

                # 先頭要素から部材名・要素番号
                elems = non.get("elements", [])
                if elems:
                    first = elems[0]  # 例 "Mg0101"
                    m = _re2.match(r'^([A-Za-z]{2,4})(\d{4})$', first)
                    if m:
                        ws32[f"{col['部材名']}{base_row + 1}"] = sym_name.get(m.group(1), "")
                        ws32[f"{col['要素番号']}{base_row + 1}"] = m.group(2)

                # 損傷の種類=NON / 損傷程度=a / 前回判定=a
                ws32[f"{col['損傷の種類']}{base_row + 2}"] = "NON"
                ws32[f"{col['損傷程度']}{base_row + 2}"] = "a"
                ws32[f"{col['前回損傷程度']}{base_row + 4}"] = "a"

                # メモ = 全要素のカンマ区切り（フロントで設定済みのmemoを優先）
                ws32[f"{col['メモ']}{base_row + 6}"] = non.get("memo", "") or ",".join(elems)

                # 写真貼り付け
                akey = non.get("photoFile", "")
                if akey and akey in photos_base64_32:
                    try:
                        import base64 as _b64n, io as _ion
                        from PIL import Image as _PILn
                        from openpyxl.drawing.image import Image as _XLImgn
                        from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor as _OCAn, AnchorMarker as _AMn
                        from openpyxl.utils.units import pixels_to_EMU as _p2en
                        b64d = photos_base64_32[akey].split(",")[-1]
                        pim = _PILn.open(_ion.BytesIO(_b64n.b64decode(b64d)))
                        th = int(min(229, 2112) * 0.95 * 0.95)
                        tw = int(th * 4 / 3)
                        pim = pim.resize((tw, th), _PILn.LANCZOS)
                        itmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                        pim.save(itmp.name, "PNG"); itmp.close()
                        xim = _XLImgn(itmp.name)
                        acol = [3, 25, 47][col_in_page]
                        arow = 11 + PAGE_ROWS * page + row_in_page * 13 + 3 - 2 + (1 if row_in_page == 0 else 0)
                        mk = _AMn(col=acol, colOff=_p2en(10), row=arow, rowOff=_p2en(15))
                        xim.anchor = _OCAn(_from=mk, ext=openpyxl.drawing.xdr.XDRPositiveSize2D(_p2en(tw), _p2en(th)))
                        ws32.add_image(xim)
                    except Exception as e:
                        print(f"NON写真貼り付けエラー: {e}")

    if "__clean_32__" in wb.sheetnames:
        del wb["__clean_32__"]
    wb.save(output_path)


def write_inspection_template(data: dict, output_path: str):
    """点検記録様式(R6)その1（1/2・2/2ヘッダー）に橋梁諸元を書き込む"""
    shutil.copy(INSPECTION_TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb["その１"]

    # 結合セルのアンカー（先頭セル）を解決して書き込む
    from openpyxl.utils import coordinate_to_tuple as _c2t, get_column_letter as _gcl
    def _anchor(cell):
        r, c = _c2t(cell)
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                return f"{_gcl(mr.min_col)}{mr.min_row}"
        return cell
    def put(cell, value, force=False):
        if not force and value in (None, ""):
            return
        ws[_anchor(cell)] = value if value is not None else ""

    # inspection_form（STEP8で編集した値）を優先、なければdataのトップレベル
    iform = data.get("inspection_form", {}) or {}
    def fv(key, fallback_key=None):
        """inspection_form優先で値を取得"""
        v = iform.get(key)
        if v not in (None, ""):
            return v
        return data.get(fallback_key or key, "")

    def fv_force(key, fallback_key=None):
        """iformにキーが存在すれば空でも返す（削除を反映するため）"""
        if key in iform:
            return iform[key], True
        v = data.get(fallback_key or key, "")
        return v, False

    # --- (1/2) 緯度経度・施設ID ---
    put("AP2", fv("起点側緯度"))
    put("AP3", fv("起点側経度"))
    put("AZ2", fv("終点側緯度"))
    put("AZ3", fv("終点側経度"))
    put("BI2", fv("施設ID"))

    # --- (1/2) ヘッダー ---
    put("H5",  fv("フリガナ"))
    put("H6",  fv("橋梁名"))
    put("AE5", fv("路線名"))
    put("AQ5", fv("管理者", "管理者"))
    put("BE5", fv("橋梁コード"))
    v_gej, f_gej = fv_force("管轄_地方整備局")
    put("AX5", v_gej, force=f_gej)
    v_jmu, f_jmu = fv_force("管轄_事務所")
    put("AX7", v_jmu, force=f_jmu)
    v_chj, f_chj = fv_force("管轄_出張所")
    put("AX8", v_chj, force=f_chj)
    put("J7",  fv("所在地"))
    put("J8",  fv("所在地至"))
    put("AG7", fv("距離標自"))
    put("AG8", fv("距離標至"))
    put("BI7", fv("調書更新年月日"))

    # --- (1/2) 橋梁諸元 ---
    kansei = "".join(str(x) for x in [
        data.get("完成年号", ""),
        data.get("完成年", "") or "",
        "年" if data.get("完成年") else "",
        data.get("完成月", "") or "",
        "月" if data.get("完成月") else "",
    ])
    put("H10", fv("供用開始日") or kansei)
    put("P10", fv("橋長"))
    put("H11", fv("上部構造形式"))
    put("H14", fv("下部構造形式"))
    put("H17", fv("基礎形式"))
    put("AU10", fv("適用示方書"))
    put("AA11", fv("全幅員"))
    put("AA12", fv("有効幅員"))
    put("AA15", fv("備考"))

    # --- (2/2) ヘッダー ---
    put("AP45", fv("起点側緯度"))
    put("AP46", fv("起点側経度"))
    put("AZ45", fv("終点側緯度"))
    put("AZ46", fv("終点側経度"))
    put("BI45", fv("施設ID"))
    put("H49", fv("フリガナ"))
    put("H50", fv("橋梁名"))
    put("AE49", fv("路線名"))
    put("AQ49", fv("管理者", "管理者"))

    # --- (2/2) 性能評価総括 ---
    PERF_ROW_CELLS = {  # 構成要素 → 開始行
        "zentai": 57, "joubu": 59, "setsuzoku": 61, "kabu": 63, "failsafe": 65, "shinshuku": 67,
    }
    PERF_COL_CELLS = {  # 評価列 → (評価セル列, 写真セル列)
        "katsu": ("N", "V"), "jishin": ("AB", "AJ"), "gou": ("AP", "AX"), "sonota": ("BD", "BL"),
    }
    for rkey, rrow in PERF_ROW_CELLS.items():
        for ckey, (vcol, pcol) in PERF_COL_CELLS.items():
            put(f"{vcol}{rrow}", iform.get(f"perf_{rkey}_{ckey}", ""))
            put(f"{pcol}{rrow}", iform.get(f"perf_{rkey}_{ckey}_photo", ""))

    # --- (2/2) 所見・診断員 ---
    put("AH54", fv("現地確認年月日"))
    put("AX54", fv("橋梁診断員"))
    put("B71",  fv("診断員所見"))

    # --- その4：現地状況写真 ---
    write_inspection_situation_photos(wb, data)

    # --- その8〜その10：STEP7データ書き込み ---
    write_inspection_step7_sheets(wb, data)

    wb.save(output_path)


def write_inspection_step7_sheets(wb, data: dict):
    """
    点検記録様式 その8〜その10 に STEP7データを書き込む。
    その8  : システム × 径間ごとにシートをコピー
    その9-1/9-2/その10: 径間ごとにシートをコピー
    写真4枚 = 1ページ(行10-49)。5枚以上は行10-49ブロックを下にコピー追加。
    """
    import copy
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl import styles as xl_styles

    inspection_items = data.get("inspection_items", [])
    inspection_evals = data.get("inspection_evals", {})
    if not inspection_items:
        return

    SYSTEM_ORDER = [
        "\u5e8a\u7248\u30fb\u5e8a\u7d44\u30b7\u30b9\u30c6\u30e0",
        "\u4e3b\u6841\u30fb\u4e3b\u69cb\u30b7\u30b9\u30c6\u30e0",
        "\u7acb\u4f53\u6a5f\u80fd\u4fdd\u6301\u30b7\u30b9\u30c6\u30e0",
        "\u652f\u70b9\u53cd\u529b\u652f\u6301\u30b7\u30b9\u30c6\u30e0",
        "\u4f4d\u7f6e\u4fdd\u6301\u30b7\u30b9\u30c6\u30e0",
        "\u652f\u70b9\u4f4d\u7f6e\u4fdd\u6301\u30b7\u30b9\u30c6\u30e0",
        "\u5730\u8868\u9762\u4f4d\u7f6e\u4fdd\u6301\u30b7\u30b9\u30c6\u30e0",
    ]
    SYSTEM_NONE = "\uff08\u30b7\u30b9\u30c6\u30e0\u672a\u5206\u985e\uff09"
    PAGE_ROWS = 40   # 行10〜49 = 40行が1ページ分
    PAGE_START = 10  # ページ本体の開始行
    PAGE_END   = 49  # ページ本体の終了行
    PER_PAGE   = 4   # 写真4枚/ページ

    # --- seqMap 再構築（その8-1→その8-2→9-1→9-2→10 の順で全体連番） ---
    SHEET_TABS = ["\u305d\u306e8-1", "\u305d\u306e8-2", "\u305d\u306e9-1", "\u305d\u306e9-2", "\u305d\u306e10"]
    seq_map = {}
    seq = 1
    def _sys_order(k):
        sys = k.split("__", 1)[1] if "__" in k else ""
        idx = SYSTEM_ORDER.index(sys) if sys in SYSTEM_ORDER else 99
        sp = int(k.split("__")[0][1:]) if k.split("__")[0][1:].isdigit() else 0
        return (idx, sp)
    for sheet in SHEET_TABS:
        sheet_items = [it for it in inspection_items if it.get("sheet") == sheet and not it.get("deleted")]
        if sheet in ("\u305d\u306e8-1", "\u305d\u306e8-2"):
            gm = {}
            for it in sheet_items:
                sp = it.get("spanNo", 1)
                sys = it.get("system") or SYSTEM_NONE
                k = f"s{sp}__{sys}"
                gm.setdefault(k, []).append(it)
            for k in sorted(gm.keys(), key=_sys_order):
                for it in sorted(gm[k], key=lambda x: x.get("order", 9999)):
                    seq_map[it["id"]] = seq
                    seq += 1
        else:
            gm = {}
            for it in sheet_items:
                sp = it.get("spanNo", 1)
                gm.setdefault(sp, []).append(it)
            for sp in sorted(gm.keys()):
                for it in sorted(gm[sp], key=lambda x: x.get("order", 9999)):
                    seq_map[it["id"]] = seq
                    seq += 1

    def write_header(ws, data, span_no=None):
        """ヘッダー共通書き込み"""
        def p(cell, val):
            if val not in (None, ""):
                c = ws[cell]
                if type(c).__name__ == "MergedCell":
                    print(f"SKIP MergedCell header: {ws.title} {cell}")
                    return
                ws[cell] = val
        p("AP2", data.get("\u8d77\u70b9\u5074\u7def\u5ea6", ""))
        p("AP3", data.get("\u8d77\u70b9\u5074\u7d4c\u5ea6", ""))
        p("AZ2", data.get("\u7d42\u70b9\u5074\u7def\u5ea6", ""))
        p("AZ3", data.get("\u7d42\u70b9\u5074\u7d4c\u5ea6", ""))
        p("BI2", data.get("\u65bd\u8a2dID", ""))
        p("H5",  data.get("\u30d5\u30ea\u30ac\u30ca", ""))
        p("H6",  data.get("\u6a4b\u6881\u540d", ""))
        p("AE5", data.get("\u8def\u7dda\u540d", ""))
        p("AQ5", data.get("\u7ba1\u7406\u8005", ""))
        p("BB5", data.get("\u6a4b\u6881\u30b3\u30fc\u30c9", ""))
        p("AC8", data.get("\u73fe\u5730\u78ba\u8a8d\u5e74\u6708\u65e5", ""))

    def copy_page_block(ws, src_start, dst_start, src_ws=None):
        """src_ws(未指定なら自身)の行src_start〜 を ws の dst_start以降にコピー（その4方式）。
        src_wsには白紙のクリーンテンプレを渡すこと（書き込み済みページを複製しない）。"""
        from openpyxl.utils import get_column_letter as _gcl
        from copy import copy as _cp
        if src_ws is None:
            src_ws = ws
        src_s = src_start
        src_e = src_start + PAGE_ROWS - 1
        num = PAGE_ROWS
        offset = dst_start - src_s

        # 結合セルの非アンカーから書式を取得するためのアンカー解決
        def get_anchor_cell(w, row, col):
            for m in w.merged_cells.ranges:
                if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
                    return w.cell(m.min_row, m.min_col)
            return None

        # 1) コピー先範囲の結合を先に解除
        for m in list(ws.merged_cells.ranges):
            if dst_start <= m.min_row <= dst_start + num - 1:
                ws.unmerge_cells(str(m))
        # 2) 書式・値コピー（srcは白紙テンプレなので値ごと運んでも残骸にならない）
        for r in range(num):
            sr, dr = src_s + r, dst_start + r
            for c in range(1, src_ws.max_column + 1):
                src_cell = src_ws.cell(sr, c)
                dst_cell = ws.cell(dr, c)
                if type(dst_cell).__name__ == "MergedCell":
                    continue
                if type(src_cell).__name__ == "MergedCell":
                    src_cell = get_anchor_cell(src_ws, sr, c) or src_cell
                if src_cell.has_style:
                    dst_cell.font          = _cp(src_cell.font)
                    dst_cell.border        = _cp(src_cell.border)
                    dst_cell.fill          = _cp(src_cell.fill)
                    dst_cell.number_format = src_cell.number_format
                    dst_cell.alignment     = _cp(src_cell.alignment)
                dst_cell.value = src_cell.value if type(src_cell).__name__ != "MergedCell" else None
            ws.row_dimensions[dr].height = src_ws.row_dimensions[sr].height
        # 3) 結合を複製
        for m in list(src_ws.merged_cells.ranges):
            if src_s <= m.min_row and m.max_row <= src_e:
                nr = f"{_gcl(m.min_col)}{m.min_row+offset}:{_gcl(m.max_col)}{m.max_row+offset}"
                try:
                    ws.merge_cells(nr)
                except Exception:
                    pass

    # 写真4枠（1ページ内）への書き込み
    # slot 0〜1: 行13-30の左右、slot 2〜3: 行31-48の左右
    # (写真番号col, 部材名col, 要素番号col, 損傷種類col, 写真行, 損傷行)
    # 左ブロック: 写真=E, 部材=J, 要素=Q, 損傷=G
    # 右ブロック: 写真=V, 部材=AA, 要素=AH, 損傷=X
    PHOTO_SLOTS = [
        ("E", "J",  "Q",  "G", 13, 14),  # slot0: 左上
        ("V", "AA", "AH", "X", 13, 14),  # slot1: 右上
        ("E", "J",  "Q",  "G", 31, 32),  # slot2: 左下
        ("V", "AA", "AH", "X", 31, 32),  # slot3: 右下
    ]
    # 写真アンカー (0始まり col, 基準row): 左上B16 / 右上S16 / 左下B34 / 右下S34
    PHOTO_ANCHORS = [
        (1, 15),   # slot0 左上 B16
        (18, 15),  # slot1 右上 S16
        (1, 33),   # slot2 左下 B34
        (18, 33),  # slot3 右下 S34
    ]

    def _normalize_key(s):
        if not s:
            return ""
        return re.sub(r"^([a-z])_+(\d)", r"\1___\2", s.lower())

    def _find_photo_key(photo_file, photos_b64):
        if not photo_file or not photos_b64:
            return None
        inp = _normalize_key(re.sub(r"\.[^.]+$", "", photo_file).strip())
        if not inp:
            return None
        if inp in photos_b64:
            return inp
        for k in photos_b64.keys():
            kn = _normalize_key(re.sub(r"\.[^.]+$", "", k))
            if kn == inp or inp in kn or kn in inp:
                return k
        return None

    def write_page(ws, items_page, page_offset, ev):
        """1ページ分(写真4枚まで)を書き込む。page_offsetは行オフセット(0, 40, 80...)"""
        base = page_offset
        photos_base64 = data.get("photos_base64", {})
        for slot_idx, (c_photo, c_member, c_elem, c_dmg, label_row, val_row) in enumerate(PHOTO_SLOTS):
            if slot_idx >= len(items_page):
                break
            it = items_page[slot_idx]
            lr = label_row + base
            vr = val_row + base
            for addr, val in [
                (f"{c_photo}{lr}", str(seq_map.get(it["id"], ""))),
                (f"{c_member}{lr}", it.get("memberName", "")),
                (f"{c_elem}{lr}", it.get("elementNoCurrent", "")),
                (f"{c_dmg}{vr}", it.get("dmgType", "")),
            ]:
                c = ws[addr]
                if type(c).__name__ == "MergedCell":
                    continue
                c.value = val
            # キャプション（状況写真参照 or 損傷写真参照）
            sit_num = it.get("sitPhotoNum", "")
            ref = it.get("refNum", "")
            if sit_num:
                ref_val = f"データ記録様式(1)状況写真{sit_num}"
            elif ref:
                ref_val = f"データ記録様式(3-2)損傷写真{ref}"
            else:
                ref_val = ""
            ref_row = (30 if label_row == 13 else 48) + base
            ref_col = 2 if c_photo == "E" else 19   # 左=B列 / 右=S列
            ws.cell(row=ref_row, column=ref_col).value = ref_val

            # 写真貼り付け（twoCellAnchor + editAs="oneCell"：印刷拡大対策）
            pkey = _find_photo_key(it.get("photoFile", ""), photos_base64)
            if pkey and pkey in photos_base64:
                try:
                    import base64 as _b64, io as _io
                    from PIL import Image as _PIL
                    from openpyxl.drawing.image import Image as _XLImg
                    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
                    from openpyxl.utils.units import pixels_to_EMU
                    b64_data = photos_base64[pkey].split(",")[-1]
                    img_bytes = _b64.b64decode(b64_data)
                    pil_img = _PIL.open(_io.BytesIO(img_bytes))
                    target_h_px = int(min(229, 2112) * 0.95 * 0.95)
                    target_w_px = int(target_h_px * 4 / 3)
                    pil_img = pil_img.resize((target_w_px, target_h_px), _PIL.LANCZOS)
                    img_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                    pil_img.save(img_tmp.name, "PNG")
                    img_tmp.close()
                    xl_img = _XLImg(img_tmp.name)
                    anchor_col, anchor_row0 = PHOTO_ANCHORS[slot_idx]
                    # from: 枠左上（解消後ファイルと同値: colOff=95250/10px, rowOff=142875/15px）
                    frm = AnchorMarker(
                        col=anchor_col, colOff=95250,
                        row=anchor_row0 + base, rowOff=142875
                    )
                    # to: from.col+16 / from.row+12（解消後ファイルの実値に一致）
                    to = AnchorMarker(
                        col=anchor_col + 16, colOff=63500,
                        row=anchor_row0 + base + 12, rowOff=127000
                    )
                    xl_img.anchor = TwoCellAnchor(editAs="oneCell", _from=frm, to=to)
                    ws.add_image(xl_img)
                except Exception as e:
                    print(f"その8-10写真貼り付けエラー: {e}")

        # 評価欄（シートによって異なる）
        # その8/9-1/9-2: ラベル行(14,19,25,27)+値行(16,21,25ラベル直下=AX25など)
        # マージセル左上に書き込む
        sheet_name = ws.title
        if sheet_name.startswith("__tmp_"):
            sheet_name = sheet_name[len("__tmp_"):]
        # 9系グレー列は出力でも空欄化（テンプレのグレー塗りと整合。保存値が古くても正しく出力）
        _gray = set()
        if sheet_name.startswith("\u305d\u306e\uff19-1"):      # その９-1
            _gray = {"katsu", "gou", "sonotaJoukyou1", "sonotaJoukyou2", "sensaku"}
        elif sheet_name.startswith("\u305d\u306e\uff19-2"):    # その９-2
            _gray = {"jishin", "gou", "sonotaJoukyou1", "sonotaJoukyou2"}
        if _gray:
            ev = {**ev}
            for _gk in _gray:
                ev[_gk] = ""
        _is_8_9 = any(sheet_name.startswith(p) for p in [
            "\u305d\u306e\uff18-1", "\u305d\u306e\uff18-2",
            "\u305d\u306e\uff19-1", "\u305d\u306e\uff19-2",
        ])
        if _is_8_9:
            r16 = 16 + base   # 想定する状況の値行
            r21 = 21 + base   # 特定事象の値行
            r25 = 25 + base   # E/M/S1/S2の値行
            r28 = 28 + base   # 所見の値行
            for addr, val in [
                (f"AJ{r16}", ev.get("katsu", "")),
                (f"AO{r16}", ev.get("jishin", "")),
                (f"AT{r16}", ev.get("gou", "")),
                (f"AY{r16}", ev.get("sonotaJoukyou1", "")),
                (f"BH{r16}", ev.get("sonotaJoukyou2", "")),
                (f"AJ{r21}", ev.get("hiro", "")),
                (f"AN{r21}", ev.get("engai", "")),
                (f"AR{r21}", ev.get("alkali", "")),
                (f"AV{r21}", ev.get("boshoku", "")),
                (f"AZ{r21}", ev.get("sensaku", "")),
                (f"BD{r21}", ev.get("shinshuku", "")),
                (f"BH{r21}", ev.get("sonotaJisho1", "")),
                (f"BM{r21}", ev.get("sonotaJisho2", "")),
                (f"AL{r25}", ev.get("kinkyu", "")),
                (f"AX{r25}", ev.get("iji", "")),
                (f"BH{r25}", ev.get("shosai", "")),
                (f"BN{r25}", ev.get("tsuiseki", "")),
                (f"AJ{r28}", ev.get("shoken", "")),
            ]:
                c = ws[addr]
                if type(c).__name__ == "MergedCell":
                    print(f"SKIP MergedCell: {ws.title} {addr}")
                    continue
                c.value = val
        else:
            # その10: E=AX14, M=AL14, S1=BF14, S2=BH14（各マージ左上）、所見=AJ17
            r14 = 14 + base
            r17 = 17 + base
            for addr, val in [
                (f"AX{r14}", ev.get("kinkyu", "")),
                (f"AL{r14}", ev.get("iji", "")),
                (f"BF{r14}", ev.get("shosai", "")),
                (f"BH{r14}", ev.get("tsuiseki", "")),
                (f"AJ{r17}", ev.get("shoken", "")),
            ]:
                c = ws[addr]
                if type(c).__name__ == "MergedCell":
                    print(f"SKIP MergedCell: {ws.title} {addr}")
                    continue
                c.value = val

    def process_sheet(template_sheet_name, groups, insert_after_name, eval_sheet=None):
        """
        グループリスト(各グループ={label, span_no, items})を受け取り、
        テンプレートシートをコピーしてシートを生成する。
        """
        tmpl_ws = wb[template_sheet_name]
        inserted_names = []

        # 汚染防止用クリーンテンプレ（非表示で保持し、各シートはここから複製）
        clean_tmpl = wb.copy_worksheet(tmpl_ws)
        clean_tmpl.title = f"__clean_{template_sheet_name}"
        clean_tmpl.sheet_state = "hidden"

        for g_idx, g in enumerate(groups):
            span_no = g["span_no"]
            items_all = g["items"]
            label = g["label"]

            if g_idx == 0:
                # 1枚目: クリーンテンプレから複製してテンプレ名にする
                ws = wb.copy_worksheet(clean_tmpl)
                ws.title = f"__tmp_{template_sheet_name}"
                ws.sheet_view.view = "pageBreakPreview"
                sheet_label = template_sheet_name
                first_ws = ws
            else:
                # 2枚目以降: クリーンテンプレから複製
                new_title = f"{template_sheet_name}({g_idx + 1})"
                wb.copy_worksheet(clean_tmpl).title = new_title
                # 挿入位置調整（前のシートの直後）
                prev_name = inserted_names[-1] if inserted_names else template_sheet_name
                prev_idx = wb.sheetnames.index(prev_name)
                new_idx = wb.sheetnames.index(new_title)
                wb.move_sheet(new_title, offset=prev_idx - new_idx + 1)
                ws = wb[new_title]
                ws.sheet_view.view = "pageBreakPreview"
                sheet_label = new_title

            inserted_names.append(sheet_label)

            # ヘッダー書き込み
            write_header(ws, data, span_no)

            # ページ分割して書き込み
            pages = [items_all[i:i+PER_PAGE] for i in range(0, len(items_all), PER_PAGE)] or [[]]
            num_pages = len(pages)
            for pi, page_items in enumerate(pages):
                page_offset = pi * PAGE_ROWS
                if pi > 0:
                    # 2ページ目以降: 白紙クリーンテンプレから行ブロックをコピー
                    src_start = PAGE_START
                    dst_start = PAGE_START + pi * PAGE_ROWS
                    copy_page_block(ws, src_start, dst_start, src_ws=clean_tmpl)
                # evalキー（9系はシート別、8系は従来の径間単位）
                eval_key = f"s{span_no}__{eval_sheet}__p{pi}" if eval_sheet else f"s{span_no}__p{pi}"
                ev = inspection_evals.get(eval_key, {})
                write_page(ws, page_items, page_offset, ev)

            # 印刷範囲・改ページを設定
            from openpyxl.utils import get_column_letter as _gcl
            from openpyxl.worksheet.pagebreak import Break as _Break
            last_row = PAGE_START + num_pages * PAGE_ROWS - 1
            ws.print_area = f"A1:BR{last_row}"
            ws.print_title_rows = "1:9"
            ws.row_breaks.brk.clear()
            for pi in range(1, num_pages):
                ws.row_breaks.append(_Break(id=PAGE_START + pi * PAGE_ROWS - 1))

        # 後処理: 元テンプレ削除 → 1枚目を正式名へ → クリーンテンプレ削除
        if groups:
            tmpl_idx = wb.sheetnames.index(template_sheet_name)
            del wb[template_sheet_name]
            first_ws.title = template_sheet_name
            cur = wb.sheetnames.index(template_sheet_name)
            wb.move_sheet(template_sheet_name, offset=tmpl_idx - cur)
        del wb[f"__clean_{template_sheet_name}"]

        return inserted_names

    # --- その8-1 / その8-2: システム × 径間 でグループ化 ---
    def _make_groups_8(sheet_key):
        # 径間単位でグループ化（各システムは径間シート内で縦に連結）
        items = [it for it in inspection_items if it.get("sheet") == sheet_key and not it.get("deleted")]
        span_map = {}
        for it in items:
            sp = it.get("spanNo", 1)
            if sp not in span_map:
                span_map[sp] = {}
            sys = it.get("system") or SYSTEM_NONE
            sys_idx = SYSTEM_ORDER.index(sys) if sys in SYSTEM_ORDER else 99
            if sys_idx not in span_map[sp]:
                span_map[sp][sys_idx] = []
            span_map[sp][sys_idx].append(it)
        groups = []
        for sp in sorted(span_map.keys()):
            merged_items = []
            for sys_idx in sorted(span_map[sp].keys()):
                sys_items = sorted(span_map[sp][sys_idx], key=lambda x: x.get("order", 9999))
                merged_items.extend(sys_items)
            groups.append({"span_no": sp, "system": "", "label": f"\u7b2c{sp}\u5f84\u9593", "items": merged_items})
        return groups

    groups_8_1 = _make_groups_8("\u305d\u306e8-1")
    if groups_8_1:
        process_sheet("\u305d\u306e\uff18-1", groups_8_1, "\u305d\u306e\uff17")
    groups_8_2 = _make_groups_8("\u305d\u306e8-2")
    if groups_8_2:
        process_sheet("\u305d\u306e\uff18-2", groups_8_2, "\u305d\u306e\uff17")

    # --- その9-1/9-2/その10: 径間でグループ化 ---
    for tmpl_name, sheet_key in [
        ("\u305d\u306e\uff19-1", "\u305d\u306e9-1"),
        ("\u305d\u306e\uff19-2", "\u305d\u306e9-2"),
        ("\u305d\u306e10",        "\u305d\u306e10"),
    ]:
        items_s = [it for it in inspection_items if it.get("sheet") == sheet_key and not it.get("deleted")]
        gm = {}
        for it in items_s:
            sp = it.get("spanNo", 1)
            gm.setdefault(sp, {"span_no": sp, "system": "", "label": f"\u7b2c{sp}\u5f84\u9593", "items": []})
            gm[sp]["items"].append(it)
        groups_s = [v for _, v in sorted(gm.items())]
        for g in groups_s:
            g["items"].sort(key=lambda x: x.get("order", 9999))
        if groups_s:
            process_sheet(tmpl_name, groups_s, None, eval_sheet=sheet_key)


def write_inspection_situation_photos(wb, data: dict):
    """点検記録様式その4（診断のための状態の把握時の現地状況写真）に写真情報を書き込む"""
    situation_photos = data.get("situation_photos", [])
    if not situation_photos:
        return

    PAGE_ROWS = 8  # 行10〜17が1ページ分
    COL_VALS = [
        {"写真番号": "H",  "径間番号": "R",  "メモ": "H"},
        {"写真番号": "AD", "径間番号": "AN", "メモ": "AD"},
        {"写真番号": "AZ", "径間番号": "BJ", "メモ": "AZ"},
    ]
    VALUE_COLS = ['H', 'AD', 'AZ', 'R', 'AN', 'BJ']

    from openpyxl.utils import column_index_from_string as c2i4
    from openpyxl.utils import get_column_letter as gcl4
    from openpyxl.worksheet.pagebreak import Break as Break4
    from copy import copy as _copy4

    def write_header_4(ws):
        ws["H5"]  = data.get("フリガナ", "")
        ws["H6"]  = data.get("橋梁名", "")
        ws["AE5"] = data.get("路線名", "")
        ws["AQ5"] = data.get("管理者", "")
        ws["AP2"] = data.get("起点側緯度", "")
        ws["AP3"] = data.get("起点側経度", "")
        ws["AZ2"] = data.get("終点側緯度", "")
        ws["AZ3"] = data.get("終点側経度", "")
        ws["BI2"] = data.get("施設ID", "")

    def copy_page_4(ws, dst_start):
        """行10〜17（1ページ分）をdst_startにコピー"""
        src_s, src_e = 10, 17
        num = src_e - src_s + 1
        for m in list(ws.merged_cells.ranges):
            if dst_start <= m.min_row <= dst_start + num - 1:
                ws.unmerge_cells(str(m))
        for r in range(num):
            sr, dr = src_s + r, dst_start + r
            for c in range(1, ws.max_column + 1):
                src = ws.cell(sr, c)
                dst = ws.cell(dr, c)
                if type(dst).__name__ == "MergedCell":
                    continue
                if src.has_style:
                    dst.font = _copy4(src.font)
                    dst.border = _copy4(src.border)
                    dst.fill = _copy4(src.fill)
                    dst.number_format = src.number_format
                    dst.alignment = _copy4(src.alignment)
                dst.value = src.value
            ws.row_dimensions[dr].height = ws.row_dimensions[sr].height
        offset = dst_start - src_s
        for m in list(ws.merged_cells.ranges):
            if src_s <= m.min_row and m.max_row <= src_e:
                nr = f"{gcl4(m.min_col)}{m.min_row+offset}:{gcl4(m.max_col)}{m.max_row+offset}"
                try:
                    ws.merge_cells(nr)
                except Exception:
                    pass
        for r in range(num):
            row = dst_start + r
            for col_letter in VALUE_COLS:
                c = ws.cell(row, c2i4(col_letter))
                if type(c).__name__ != 'MergedCell':
                    c.value = None

    base_sheet_4 = "その４"
    span_nos_4 = sorted(set(p.get("spanNo") for p in situation_photos if p.get("spanNo") is not None))
    if not span_nos_4:
        span_nos_4 = [1]
    insert_pos_4 = wb.sheetnames.index(base_sheet_4)

    ws4_sheets = {}
    for s_idx, span_no in enumerate(span_nos_4):
        if s_idx == 0:
            ws4_sheets[span_no] = wb[base_sheet_4]
        else:
            label = f"{base_sheet_4} ({s_idx + 1})"
            if label in wb.sheetnames:
                del wb[label]
            ws_new = wb.copy_worksheet(wb[base_sheet_4])
            ws_new.title = label
            target_idx = insert_pos_4 + s_idx
            current_idx = wb.sheetnames.index(label)
            wb.move_sheet(label, offset=target_idx - current_idx)
            ws_new.sheet_view.showGridLines = False
            ws_new.print_title_rows = "$1:$9"
            ws4_sheets[span_no] = ws_new

    photos_base64 = data.get("photos_base64", {})
    for s_idx, span_no in enumerate(span_nos_4):
        ws = ws4_sheets[span_no]
        span_photos = [p for p in situation_photos if p.get("spanNo") == span_no]
        pages = max(1, -(-len(span_photos) // 6))

        write_header_4(ws)
        ws.sheet_view.view = "pageBreakPreview"

        # 印刷設定をデータ記録様式と同じに統一（fitToPage無効、scale=95%）
        from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
        if ws.sheet_properties is None:
            ws.sheet_properties = WorksheetProperties()
        if ws.sheet_properties.pageSetUpPr is None:
            ws.sheet_properties.pageSetUpPr = PageSetupProperties()
        ws.sheet_properties.pageSetUpPr.fitToPage = False
        ws.page_setup.fitToWidth = 0
        ws.page_setup.fitToHeight = 0
        ws.page_setup.scale = 95
        ws.page_margins.top = 0.51181102362204722
        ws.page_margins.header = 0.31496062992125984
        ws.page_margins.footer = 0.31496062992125984

        for row in range(10, 18):
            for col_letter in VALUE_COLS:
                c = ws.cell(row, c2i4(col_letter))
                if type(c).__name__ != 'MergedCell':
                    c.value = None

        for page in range(1, pages):
            copy_page_4(ws, 10 + PAGE_ROWS * page)

        for page in range(1, pages):
            ws.row_breaks.append(Break4(id=10 + PAGE_ROWS * page - 1))
        ws.print_area = f"A1:{gcl4(ws.max_column)}{9 + PAGE_ROWS * pages}"

        for i, photo in enumerate(span_photos):
            page = i // 6
            pos = i % 6
            col = COL_VALS[pos % 3]
            base = 10 + PAGE_ROWS * page + (pos // 3) * 4
            ws[f"{col['写真番号']}{base}"] = photo.get("photoNum", "")
            ws[f"{col['径間番号']}{base}"] = photo.get("spanNo", "")
            ws[f"{col['メモ']}{base + 1}"] = photo.get("memo", "")

            assigned_key = photo.get("assignedKey", "")
            if assigned_key and assigned_key in photos_base64:
                try:
                    import base64 as _b64, io as _io
                    from PIL import Image as _PIL
                    from openpyxl.drawing.image import Image as _XLImg
                    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                    from openpyxl.utils.units import pixels_to_EMU
                    b64_data = photos_base64[assigned_key].split(",")[-1]
                    img_bytes = _b64.b64decode(b64_data)
                    pil_img = _PIL.open(_io.BytesIO(img_bytes))
                    target_h_px = int(228 * 0.95)
                    target_w_px = int(target_h_px * 4 / 3)
                    pil_img = pil_img.resize((target_w_px, target_h_px), _PIL.LANCZOS)
                    img_tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                    pil_img.save(img_tmp.name, "PNG")
                    img_tmp.close()
                    xl_img = _XLImg(img_tmp.name)
                    row_offset_px = int((228 - target_h_px) / 2)
                    anchor_cols_idx = [5, 27, 49]  # F=5, AB=27, AX=49 (0-indexed)
                    col_idx = anchor_cols_idx[pos % 3]
                    row_idx = base + 2
                    # from は元のOneCellAnchorと同値。to は手動修正後ファイル(Excel)の実値を移植。
                    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor
                    co = pixels_to_EMU(int((2112 - target_w_px) / 2))
                    ro = pixels_to_EMU(row_offset_px)
                    frm = AnchorMarker(col=col_idx, colOff=co, row=row_idx, rowOff=ro)
                    to = AnchorMarker(col=col_idx + 18, colOff=99579, row=row_idx, rowOff=2114550)
                    xl_img.anchor = TwoCellAnchor(editAs="oneCell", _from=frm, to=to)
                    ws.add_image(xl_img)
                except Exception as e:
                    print(f"その4写真貼り付けエラー: {e}")


@app.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=_upload_suffix(file)) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        data = read_excel(tmp_path)
        return {"status": "ok", "data": data}
    finally:
        os.unlink(tmp_path)

@app.post("/extract-members")
async def extract_members(file: UploadFile = File(...)):
    """
    点検調書 (.xls) から径間別部材リストを返す。
    判別：B2セルでその１１/その１２を判定（シート名の表記ゆれに非依存）
    径間：Z2セルの値を径間番号として使用
    重複排除：工種・材料・部材名称・記号・要素番号の組み合わせで集約
    """
    with tempfile.NamedTemporaryFile(delete=False, suffix=_upload_suffix(file)) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        wb = open_workbook_any(tmp_path)
        spans = {}

        for sname in wb.sheet_names():
            sh = wb.sheet_by_name(sname)
            if sh.nrows < 10:
                continue
            b2 = str(sh.cell_value(1, 1)).strip()
            z2 = sh.cell_value(1, 25)
            if "データ記録様式（その１１）" in b2:
                sheet_type = "major"
            elif "データ記録様式（その１２）" in b2:
                sheet_type = "other"
            else:
                continue
            if not z2:
                continue
            try:
                span_no = int(float(z2))
            except (ValueError, TypeError):
                continue
            if span_no not in spans:
                spans[span_no] = {"major": [], "other": []}
            seen = set()
            for r in range(9, sh.nrows):
                koshu   = str(sh.cell_value(r, 1)).strip()
                zairyo  = str(sh.cell_value(r, 4)).strip()
                name    = str(sh.cell_value(r, 7)).strip()
                symbol  = str(sh.cell_value(r, 16)).strip()
                elem_no = str(sh.cell_value(r, 19)).strip()
                if not name or name in ("", "nan"):
                    continue
                key = (koshu, zairyo, name, symbol, elem_no)
                if key in seen:
                    continue
                seen.add(key)
                spans[span_no][sheet_type].append({
                    "koshu": koshu, "zairyo": zairyo, "name": name,
                    "symbol": symbol, "element_no": elem_no,
                })

        result = [
            {"span_no": no, "major": spans[no]["major"], "other": spans[no]["other"]}
            for no in sorted(spans.keys())
        ]
        prev_summary = extract_prev_summary_34(tmp_path)
        return {"spans": result, "prev_summary": prev_summary}
    finally:
        os.unlink(tmp_path)


def extract_situation_photos(xls_path: str):
    wb = open_workbook_any(xls_path)
    photos = []
    for sheet_name in wb.sheet_names():
        if not (sheet_name == "その３" or sheet_name.startswith("その３ ")):
            continue
        try:
            ws = wb.sheet_by_name(sheet_name)
        except Exception:
            continue
        span_val = ws.cell_value(1, 25)
        try:
            span_no = int(float(span_val))
        except (ValueError, TypeError):
            span_no = 1
        col_pairs = [(3, 7), (22, 26), (41, 45)]
        r = 9
        while r < ws.nrows:
            label0 = str(ws.cell_value(r, col_pairs[0][0])).strip()
            if label0 == "写真番号":
                for lc, vc in col_pairs:
                    num_val = ws.cell_value(r, vc)
                    if not num_val and num_val != 0:
                        continue
                    try:
                        photo_num = int(float(num_val))
                    except (ValueError, TypeError):
                        continue
                    if photo_num == 0:
                        continue
                    span_row_val = ws.cell_value(r + 1, vc) if r + 1 < ws.nrows else ""
                    try:
                        card_span = int(float(span_row_val)) if span_row_val else span_no
                    except (ValueError, TypeError):
                        card_span = span_no
                    memo_val = ws.cell_value(r + 2, vc) if r + 2 < ws.nrows else ""
                    memo = str(memo_val).strip() if memo_val else ""
                    photos.append({
                        "photoNum": photo_num,
                        "spanNo": card_span,
                        "memo": memo,
                        "assignedKey": "",
                    })
                r += 3
            else:
                r += 1
    photos.sort(key=lambda x: (x["spanNo"], x["photoNum"]))
    return photos


@app.post("/extract-situation-photos")
async def api_extract_situation_photos(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=_upload_suffix(file)) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        photos = extract_situation_photos(tmp_path)
        return {"status": "ok", "situation_photos": photos}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def extract_situation_photos_list(xls_path: str):
    """現地状況写真リスト.xlsx（A=写真番号,B=径間,C=メモ欄,D=写真ファイル）を読み取る。
    B列(径間)が整数化できる行のみ採用し、予備行・プレースホルダ行は除外する。"""
    wb = open_workbook_any(xls_path)
    ws = wb.sheet_by_name(wb.sheet_names()[0])
    photos = []
    for r in range(1, ws.nrows):
        span_val = ws.cell_value(r, 1)
        try:
            span_no = int(float(span_val))
        except (ValueError, TypeError):
            continue
        num_val = ws.cell_value(r, 0)
        try:
            photo_num = int(float(num_val))
        except (ValueError, TypeError):
            photo_num = r
        memo_val = ws.cell_value(r, 2)
        memo = str(memo_val).strip() if memo_val not in (None, "") else ""
        file_val = ws.cell_value(r, 3)
        file_raw = str(file_val).strip() if file_val not in (None, "") else ""
        photos.append({
            "photoNum": photo_num,
            "spanNo": span_no,
            "memo": memo,
            "photoFileRaw": file_raw,
        })
    photos.sort(key=lambda x: (x["spanNo"], x["photoNum"]))
    return photos


@app.post("/extract-situation-photos-list")
async def api_extract_situation_photos_list(file: UploadFile = File(...)):
    with tempfile.NamedTemporaryFile(delete=False, suffix=_upload_suffix(file)) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        photos = extract_situation_photos_list(tmp_path)
        return {"status": "ok", "situation_photos": photos}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@app.post("/export")
async def export(file: UploadFile = File(...), bridge_data: str = Form(None)):
    """
    bridge_data: JSON文字列（App.jsxのbridgeData、members_by_spanを含む）
    fileが送られた場合はparse-excelも実行して統合
    """

    with tempfile.NamedTemporaryFile(delete=False, suffix=_upload_suffix(file)) as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    out_path = tmp_path + "_out.xlsx"
    try:
        # XLSから橋梁情報を読み込む
        data = read_excel(tmp_path)

        # bridge_dataが渡されていればmembers_by_spanを統合
        if bridge_data:
            try:
                bd = json.loads(bridge_data)
                data["members_by_span"] = bd.get("members_by_span", [])
            except Exception:
                pass

        write_template(data, out_path)
        bridge_name = data["橋梁名"]
        return FileResponse(
            out_path,
            filename=f"{bridge_name}_データ記録様式.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    finally:
        os.unlink(tmp_path)


@app.post("/export-from-data")
async def export_from_data(request: Request):
    """
    XLSファイル不要。bridge_dataのJSONだけでテンプレートに書き込んで出力。
    """
    try:
        bd = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="bridge_dataのJSONが不正です")

    # 橋梁情報のデフォルト値を補完
    data = {
        "橋梁名":       bd.get("橋梁名", ""),
        "フリガナ":     bd.get("フリガナ", ""),
        "路線名":       bd.get("路線名", ""),
        "管理者":       bd.get("管理者", ""),
        "所在地":       bd.get("所在地", ""),
        "橋長":         bd.get("橋長", ""),
        "径間数":       bd.get("径間数", 0),
        "上部構造形式": bd.get("上部構造形式", ""),
        "完成年号":     bd.get("完成年号", ""),
        "完成年":       bd.get("完成年", 0),
        "完成月":       bd.get("完成月", 0),
        "members_by_span":    bd.get("members_by_span", []),
        "damages_data":       bd.get("damages_data", []),
        "non_photos":         bd.get("non_photos", []),
        "memo_templates":     bd.get("memo_templates", {}),
        "situation_photos":   bd.get("situation_photos", []),
        "photos_base64":      bd.get("photos_base64", {}),
        "prev_summary":       bd.get("prev_summary", {}),
        "inspection_form":    bd.get("inspection_form", {}),
    }

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        out_path = tmp.name

    try:
        write_template(data, out_path)
        bridge_name = data["橋梁名"] or "調書"
        return FileResponse(
            out_path,
            filename=f"{bridge_name}_データ記録様式.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        os.unlink(out_path)
        raise HTTPException(status_code=500, detail=f"出力エラー: {str(e)}")


@app.post("/export-inspection-format")
async def export_inspection_format(request: Request):
    """
    点検記録様式(R6)の出力。
    その1（橋梁諸元・1/2、2/2ヘッダー）に橋梁情報を転記して返す。
    """
    try:
        bd = await request.json()
    except Exception:
        bd = {}

    bridge_name = bd.get("橋梁名") or "調書"

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        out_path = tmp.name

    try:
        write_inspection_template(bd, out_path)
        return FileResponse(
            out_path,
            filename=f"{bridge_name}_点検記録様式.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        os.unlink(out_path)
        raise HTTPException(status_code=500, detail=f"出力エラー: {traceback.format_exc()}")


def _chosho_perf_cells(bd):
    """様式-1の性能評価(活荷重H/地震S/豪雨AD)と写真番号(N/Y/AJ、評価の1行下)を、
    STEP7評価+チェック済み写真(オレンジ採番)から算出して返す。
    評価=最悪値集約(C>B>A)。写真番号は評価がB/Cの構成要素×状況のみ、
    その評価を生んだページ内で最初にチェック済み(chosho)のアイテムのオレンジ番号を採用。
    橋全体(zentai)は評価のみ(inspection_formの手動値)、写真番号は対象外。"""
    SHEET_TABS = ["その8-1", "その8-2", "その9-1", "その9-2", "その10"]
    SYSTEM_ORDER = ["床版・床組システム", "主桁・主構システム", "立体機能保持システム",
                    "支点反力支持システム", "位置保持システム", "支点位置保持システム", "地表面位置保持システム"]
    SYSTEM_NONE = "（システム未分類）"
    KOSEI_OF_SYSTEM = {
        "床版・床組システム": "joubu", "主桁・主構システム": "joubu", "立体機能保持システム": "joubu",
        "支点位置保持システム": "kabu", "地表面位置保持システム": "kabu",
        "支点反力支持システム": "setsuzoku",
        "その9-1": "failsafe", "その9-2": "shinshuku",
    }
    ROW = {"joubu": 24, "setsuzoku": 26, "kabu": 28, "failsafe": 30, "shinshuku": 32}
    COL = {"katsu": "H", "jishin": "S", "gou": "AD"}
    PHOTO_ROW = {"joubu": 25, "setsuzoku": 27, "kabu": 29, "failsafe": 31, "shinshuku": 33}
    PHOTO_COL = {"katsu": "N", "jishin": "Y", "gou": "AJ"}
    PER_BLOCK = 4  # InspectionPhotos.jsxと同じ(写真2x2)
    rank = {"A": 1, "B": 2, "C": 3}

    items = bd.get("inspection_items", []) or []
    evals = bd.get("inspection_evals", {}) or {}

    # オレンジ採番(チェック済み・橋全体通し・最大20)を様式-2と同じ順序で再現
    orange_of_id = {}
    ordered = []
    for sheet in SHEET_TABS:
        sheet_items = [it for it in items if it.get("sheet") == sheet and not it.get("deleted")]
        gm = {}
        for it in sheet_items:
            sp = it.get("spanNo", 1) or 1
            sys = it.get("system") or SYSTEM_NONE
            key = f"s{sp}__{sys}" if sheet in ("その8-1", "その8-2") else f"s{sp}"
            gm.setdefault(key, {"spanNo": sp, "system": sys, "items": []})["items"].append(it)
        def gksort(g):
            ai = SYSTEM_ORDER.index(g["system"]) if g["system"] in SYSTEM_ORDER else 99
            return (ai, g["spanNo"])
        for g in sorted(gm.values(), key=gksort):
            for it in sorted(g["items"], key=lambda x: x.get("order", 9999) if x.get("order") is not None else 9999):
                ordered.append((sheet, it))
    orange = 0
    for sheet, it in ordered:
        if not it.get("chosho"):
            continue
        orange += 1
        if orange > 20:
            break
        orange_of_id[it.get("id")] = orange

    # グループ化キー(s{sp}__{system} or s{sp})ごとにアイテムをPER_BLOCKずつページ分割
    # → evalsキー "{gkey}__p{pi}" に対応する pageItems を作る
    page_items_of_evalkey = {}
    gm2 = {}
    for sheet in SHEET_TABS:
        sheet_items = [it for it in items if it.get("sheet") == sheet and not it.get("deleted")]
        for it in sheet_items:
            sp = it.get("spanNo", 1) or 1
            sys = it.get("system") or SYSTEM_NONE
            key = f"s{sp}__{sys}" if sheet in ("その8-1", "その8-2") else f"s{sp}__{sheet}"
            gm2.setdefault(key, []).append(it)
    for gkey, glist in gm2.items():
        glist_sorted = sorted(glist, key=lambda x: x.get("order", 9999) if x.get("order") is not None else 9999)
        for i in range(0, max(len(glist_sorted), 1), PER_BLOCK):
            pi = i // PER_BLOCK
            page_items_of_evalkey[f"{gkey}__p{pi}"] = glist_sorted[i:i + PER_BLOCK]

    # 構成要素ごとの状況別 最悪値 と その評価キー(どのページか)
    agg = {}        # kosei -> jk -> 値
    agg_key = {}    # kosei -> jk -> evalsキー(最悪値を出したページ)
    pat = re.compile(r"^s\d+__(.+)__p\d+$")
    for key, ev in evals.items():
        m = pat.match(key)
        if not m:
            continue
        kosei = KOSEI_OF_SYSTEM.get(m.group(1))
        if not kosei:
            continue
        d = agg.setdefault(kosei, {})
        dk = agg_key.setdefault(kosei, {})
        for jk in ("katsu", "jishin", "gou"):
            v = (ev or {}).get(jk, "")
            if rank.get(v, 0) > rank.get(d.get(jk, ""), 0):
                d[jk] = v
                dk[jk] = key

    cells = {}
    for kosei in ROW.keys():
        jv = agg.get(kosei, {})
        photo_nums = {"katsu": [], "jishin": [], "gou": []}
        any_bc = any(jv.get(jk, "") in ("B", "C") for jk in ("katsu", "jishin", "gou"))
        any_a = any(jv.get(jk, "") == "A" for jk in ("katsu", "jishin", "gou"))

        # 状況(jk)ごとに、該当する全ページを走査してB/Cページの写真を全件集約
        for evalkey, page_items in page_items_of_evalkey.items():
            m2 = pat.match(evalkey)
            if not m2 or KOSEI_OF_SYSTEM.get(m2.group(1)) != kosei:
                continue
            ev = evals.get(evalkey, {}) or {}
            for jk in ("katsu", "jishin", "gou"):
                if ev.get(jk, "") in ("B", "C"):
                    for it in page_items:
                        if it.get("chosho") and it.get("id") in orange_of_id:
                            n = orange_of_id[it.get("id")]
                            if n not in photo_nums[jk]:
                                photo_nums[jk].append(n)

        for jk, col in COL.items():
            val = jv.get(jk, "")
            if val:
                cells[f"{col}{ROW[kosei]}"] = val
            nums = sorted(photo_nums[jk])
            if nums:
                cells[f"{PHOTO_COL[jk]}{PHOTO_ROW[kosei]}"] = ",".join(str(n) for n in nums)

        # B/Cが1つも無く、Aが1つ以上ある構成要素は、配下の全写真を活荷重列に集約
        if not any_bc and any_a:
            all_nums = []
            for evalkey, page_items in page_items_of_evalkey.items():
                m2 = pat.match(evalkey)
                if not m2 or KOSEI_OF_SYSTEM.get(m2.group(1)) != kosei:
                    continue
                for it in page_items:
                    if it.get("chosho") and it.get("id") in orange_of_id:
                        n = orange_of_id[it.get("id")]
                        if n not in all_nums:
                            all_nums.append(n)
            all_nums.sort()
            if all_nums:
                cells[f"{PHOTO_COL['katsu']}{PHOTO_ROW[kosei]}"] = ",".join(str(n) for n in all_nums)

    # 橋全体(zentai)は評価のみ手動値(写真番号は対象外)
    form = bd.get("inspection_form", {}) or {}
    for jk, col in COL.items():
        val = str(form.get(f"perf_zentai_{jk}", "") or "")
        if val:
            cells[f"{col}22"] = val
    return cells


def _chosho_form2_cells(bd):
    """様式-2の文字メタ(構成要素/想定する状況/構成要素の状態/写真番号)を生成。
    オレンジ採番(チェック済み・橋全体通し・最大20)をフロントと同じ順序で再現。
    状態/状況は構成要素ごとの最悪値(C>B>A、同点は活荷重→地震→豪雨→その他の先頭優先)。"""
    SHEET_TABS = ["その8-1", "その8-2", "その9-1", "その9-2", "その10"]
    SYSTEM_ORDER = ["床版・床組システム", "主桁・主構システム", "立体機能保持システム",
                    "支点反力支持システム", "位置保持システム", "支点位置保持システム", "地表面位置保持システム"]
    SYSTEM_NONE = "（システム未分類）"
    KOSEI_OF_SYSTEM = {
        "床版・床組システム": "joubu", "主桁・主構システム": "joubu", "立体機能保持システム": "joubu",
        "支点位置保持システム": "kabu", "地表面位置保持システム": "kabu",
        "支点反力支持システム": "setsuzoku",
        "その9-1": "failsafe", "その9-2": "shinshuku",
    }
    KOSEI_LABEL = {
        "joubu": "上部構造", "kabu": "下部構造", "setsuzoku": "上下部接続部",
        "failsafe": "その他（フェールセーフ）", "shinshuku": "その他（伸縮装置）",
    }
    SIT_ORDER = ["katsu", "jishin", "gou", "sonota"]
    SIT_LABEL = {"katsu": "1-活荷重", "jishin": "2-地震", "gou": "3-豪雨･出水", "sonota": "4-その他"}
    rank = {"A": 1, "B": 2, "C": 3}

    items = bd.get("inspection_items", []) or []
    evals = bd.get("inspection_evals", {}) or {}

    # 構成要素ごとの (状況→最悪値)
    agg = {}
    pat = re.compile(r"^s\d+__(.+)__p\d+$")
    sit_src = {"katsu": ["katsu"], "jishin": ["jishin"], "gou": ["gou"],
               "sonota": ["sonotaJoukyou1", "sonotaJoukyou2"]}
    for key, ev in evals.items():
        m = pat.match(key)
        if not m:
            continue
        kosei = KOSEI_OF_SYSTEM.get(m.group(1))
        if not kosei:
            continue
        d = agg.setdefault(kosei, {})
        for sit, srcs in sit_src.items():
            for ek in srcs:
                v = (ev or {}).get(ek, "")
                if rank.get(v, 0) > rank.get(d.get(sit, ""), 0):
                    d[sit] = v
    # 構成要素ごとの (最悪値, 最悪状況ラベル)
    kosei_state = {}
    for kosei, d in agg.items():
        worst = ""
        for sit in SIT_ORDER:
            if rank.get(d.get(sit, ""), 0) > rank.get(worst, 0):
                worst = d.get(sit, "")
        if not worst:
            continue
        sit = next((s for s in SIT_ORDER if d.get(s, "") == worst), "")
        kosei_state[kosei] = (worst, SIT_LABEL.get(sit, ""))

    # オレンジ採番順の再現
    ordered = []
    for sheet in SHEET_TABS:
        sheet_items = [it for it in items if it.get("sheet") == sheet and not it.get("deleted")]
        gm = {}
        for it in sheet_items:
            sp = it.get("spanNo", 1) or 1
            sys = it.get("system") or SYSTEM_NONE
            key = f"s{sp}__{sys}" if sheet in ("その8-1", "その8-2") else f"s{sp}"
            gm.setdefault(key, {"spanNo": sp, "system": sys, "items": []})["items"].append(it)
        def gksort(g):
            ai = SYSTEM_ORDER.index(g["system"]) if g["system"] in SYSTEM_ORDER else 99
            return (ai, g["spanNo"])
        for g in sorted(gm.values(), key=gksort):
            for it in sorted(g["items"], key=lambda x: x.get("order", 9999) if x.get("order") is not None else 9999):
                ordered.append((sheet, it))

    text_cells = {}
    num_cells = {}
    orange_list = []
    orange = 0
    for sheet, it in ordered:
        if not it.get("chosho"):
            continue
        orange += 1
        if orange > 20:
            break
        N = orange
        orange_list.append((N, it))
        pair = (N - 1) // 2
        left = (N % 2 == 1)
        cr = 7 + 19 * pair
        sr = 8 + 19 * pair
        pr = 24 + 19 * pair
        if left:
            c_kosei, c_sit, c_state, c_num = f"G{cr}", f"D{sr}", f"M{sr}", f"C{pr}"
        else:
            c_kosei, c_sit, c_state, c_num = f"V{cr}", f"S{sr}", f"AB{sr}", f"R{pr}"
        if sheet in ("その8-1", "その8-2"):
            kosei = KOSEI_OF_SYSTEM.get(it.get("system") or "")
        elif sheet == "その9-1":
            kosei = "failsafe"
        elif sheet == "その9-2":
            kosei = "shinshuku"
        else:
            kosei = None
        if kosei and kosei in KOSEI_LABEL:
            text_cells[c_kosei] = KOSEI_LABEL[kosei]
            st = kosei_state.get(kosei)
            if st:
                text_cells[c_state] = st[0]
                text_cells[c_sit] = st[1]
        elif sheet == "その10":
            text_cells[c_kosei] = "その他"
        num_cells[c_num] = N
    n_orange = min(sum(1 for _s, _it in ordered if _it.get("chosho")), 20)
    return text_cells, num_cells, n_orange, orange_list


def _chosho_zenkei_part(zen_durl):
    """様式-1の全景写真(1枚)を貼るパーツを生成。
    参考xlsxの実測アンカー(twoCellAnchor editAs=oneCell)を初期値に使用。
    戻り値: (media[(zipname,bytes)], drawing_xml, drawing_rels_xml, ext)。"""
    import base64 as _b64
    if not zen_durl or "," not in zen_durl:
        return [], "", "", ""
    head, b64 = zen_durl.split(",", 1)
    ext = "jpeg" if ("jpeg" in head or "jpg" in head) else "png"
    try:
        raw = _b64.b64decode(b64)
    except Exception:
        return [], "", "", ""
    IMG_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
    fname = "image_zenkei.%s" % ext
    media = [("xl/media/%s" % fname, raw)]
    anchor = (
        '<xdr:twoCellAnchor editAs="oneCell">'
        '<xdr:from><xdr:col>17</xdr:col><xdr:colOff>194812</xdr:colOff><xdr:row>36</xdr:row><xdr:rowOff>57151</xdr:rowOff></xdr:from>'
        '<xdr:to><xdr:col>31</xdr:col><xdr:colOff>24266</xdr:colOff><xdr:row>47</xdr:row><xdr:rowOff>131174</xdr:rowOff></xdr:to>'
        '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="1" name="全景写真"/><xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr></xdr:nvPicPr>'
        '<xdr:blipFill><a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:embed="rId1"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
        '<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="2896504" cy="2169523"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
        '</xdr:pic><xdr:clientData/></xdr:twoCellAnchor>'
    )
    drawing_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
                   '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                   'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
                   + anchor + '</xdr:wsDr>')
    drawing_rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    '<Relationship Id="rId1" Type="%s" Target="../media/%s"/></Relationships>' % (IMG_REL, fname))
    return media, drawing_xml, drawing_rels, ext


def _chosho_image_parts(orange_list, photos_base64):
    """オレンジ写真を様式-2へ貼るパーツを生成(縦横比保持・高さ最大)。
    戻り値: (media[(zipname,bytes)], drawing_xml, drawing_rels_xml, exts(set))
    保護・x14を壊さないようopenpyxlは使わずXML/zipを直接構築する。"""
    import base64 as _b64
    from io import BytesIO
    try:
        from PIL import Image as _PILImage
    except Exception:
        _PILImage = None
    BLOCK_H = 2228850  # 写真ブロック高さ(行10-22 = 13行×13.5pt) EMU
    IMG_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"
    media = []
    anchors = []
    rels = []
    exts = set()
    rid = 0
    for (N, it) in orange_list:
        key = it.get("choshoKey") or ""
        durl = (photos_base64 or {}).get(key, "")
        if not durl or "," not in durl:
            continue
        head, b64 = durl.split(",", 1)
        ext = "jpeg" if ("jpeg" in head or "jpg" in head) else "png"
        try:
            raw = _b64.b64decode(b64)
        except Exception:
            continue
        # 縦横比から幅を算出(高さ=ブロック高さ最大)
        cy = BLOCK_H
        cx = int(BLOCK_H * 4 / 3)  # 既定4:3
        if _PILImage is not None:
            try:
                with _PILImage.open(BytesIO(raw)) as im:
                    w, h = im.size
                    if h > 0:
                        cx = int(BLOCK_H * w / h)
            except Exception:
                pass
        rid += 1
        fname = "image_c%d.%s" % (N, ext)
        media.append(("xl/media/%s" % fname, raw))
        exts.add(ext)
        rels.append('<Relationship Id="rId%d" Type="%s" Target="../media/%s"/>' % (rid, IMG_REL, fname))
        pair = (N - 1) // 2
        left = (N % 2 == 1)
        fc = 3 if left else 18   # 半列左へ: 左=D(3)/右=S(18) +colOff
        coff = 185737            # 半列幅(width4.875≈371475 EMUの半分)
        fr = 9 + 19 * pair
        name = "写真%02d" % N
        anchors.append(
            '<xdr:oneCellAnchor>'
            '<xdr:from><xdr:col>%d</xdr:col><xdr:colOff>%d</xdr:colOff><xdr:row>%d</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
            '<xdr:ext cx="%d" cy="%d"/>'
            '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="%d" name="%s"/><xdr:cNvPicPr><a:picLocks noChangeAspect="1"/></xdr:cNvPicPr></xdr:nvPicPr>'
            '<xdr:blipFill><a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:embed="rId%d"/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
            '<xdr:spPr><a:xfrm><a:off x="0" y="0"/><a:ext cx="%d" cy="%d"/></a:xfrm><a:prstGeom prst="rect"><a:avLst/></a:prstGeom></xdr:spPr>'
            '</xdr:pic><xdr:clientData/></xdr:oneCellAnchor>'
            % (fc, coff, fr, cx, cy, N, name, rid, cx, cy)
        )
    if not anchors:
        return [], "", "", set()
    drawing_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
                   '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
                   'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
                   + "".join(anchors) + '</xdr:wsDr>')
    drawing_rels = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n'
                    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                    + "".join(rels) + '</Relationships>')
    return media, drawing_xml, drawing_rels, exts


def write_chosho(bd: dict, out_path: str):
    """道路橋記録様式(R6)を出力。様式-1(諸元/性能評価)・様式-2(状況写真の文字メタ)を
    テンプレのxmlを直接パッチして転記する。入力規則(x14)・シート保護等を壊さず全要素を保持。"""
    _iform1 = bd.get("inspection_form", {}) or {}
    def _f1(key, fallback_key=None):
        if key in _iform1 and str(_iform1.get(key, "")).strip() != "":
            return _iform1.get(key)
        return bd.get(fallback_key or key, "")
    cells1 = {
        "A6": str(bd.get("橋梁名", "") or ""),    # 橋梁名
        "D7": str(bd.get("フリガナ", "") or ""),  # フリガナ
        "O6": str(bd.get("路線名", "") or ""),    # 路線名
        "AA6": str(_f1("所在地") or ""),           # 所在地
        "AM4": str(_f1("起点側緯度") or ""),       # 緯度(起点側)
        "AM5": str(_f1("起点側経度") or ""),       # 経度(起点側)
        "AS5": str(_f1("施設ID") or ""),           # 施設ID
        "A9":  str(_f1("管理者") or ""),           # 管理者名
        "V9":  str(_f1("路下条件") or ""),         # 路下条件
        "AF9": str(_f1("代替路有無") or ""),       # 代替路の有無
        "AK9": str(_f1("自専道一般道") or ""),     # 自専道or一般道
        "AQ9": str(_f1("緊急輸送道路") or ""),     # 緊急輸送道路
        "AV9": str(_f1("占用物件") or ""),         # 占用物件(名称)
        "A14": str(_f1("健全性区分") or ""),       # 告示に基づく健全性の診断の区分
        "S14": str(_f1("橋長") or ""),             # 橋長
        "Y18": str(_f1("現地確認年月日") or ""),   # 定期点検実施年月日
        "AN18": str(_f1("診断員会社名") or ""),    # 定期点検者(会社名)
        "AV18": str(_f1("診断員氏名") or ""),      # 定期点検者(氏名)
    }
    # 架設年度(O14): STEP8「供用開始日」の文字列を解析。
    # 複数行(改行区切り)入力時は最初に判定できた行を採用。
    # 1)「不明」を含む→不明 2)西暦4桁を含む→抽出 3)元号+和暦年を含む→西暦変換 4)該当無し→不明
    def _wareki_to_seireki(gengo, year):
        try:
            y = int(str(year).strip())
        except (TypeError, ValueError):
            return ""
        base = {"明治": 1867, "大正": 1911, "昭和": 1925, "平成": 1988, "令和": 2018}.get(str(gengo))
        if base is None or y <= 0:
            return ""
        return str(base + y)
    def _parse_kyoyo_line(line):
        t = str(line or "").strip()
        if not t:
            return None
        if "不明" in t:
            return "不明"
        m = re.search(r"(18|19|20)\d{2}", t)
        if m:
            return m.group(0)
        m = re.search(r"(明治|大正|昭和|平成|令和)\s*(\d{1,2})\s*年?", t)
        if m:
            seireki = _wareki_to_seireki(m.group(1), m.group(2))
            if seireki:
                return seireki
        return None
    def _kyoyo_value():
        raw = str(_f1("供用開始日") or "")
        for line in raw.splitlines():
            v = _parse_kyoyo_line(line)
            if v:
                return v
        return "不明"
    cells1["O14"] = _kyoyo_value()
    # 橋梁形式(様式-1用・構造形式一覧から選択)
    cells1["AA15"] = str(_f1("様式1_上部構造形式") or "")
    cells1["AJ15"] = str(_f1("様式1_下部構造形式") or "")
    cells1["AR15"] = str(_f1("様式1_基礎形式") or "")
    # 幅員(W14): 有効幅員を小数点以下第2位を四捨五入して第1位まで記入
    def _round1(v):
        try:
            from decimal import Decimal, ROUND_HALF_UP
            d = Decimal(str(v)).quantize(Decimal("0.1"), rounding=ROUND_HALF_UP)
            return format(d, "f")
        except Exception:
            return str(v) if v not in (None, "") else ""
    cells1["W14"] = _round1(_f1("有効幅員"))
    cells1.update(_chosho_perf_cells(bd))   # 様式-1 性能評価(活荷重/地震/豪雨)
    text2, num2, n_orange, orange_list = _chosho_form2_cells(bd)   # 様式-2 文字メタ + 写真番号 + 枚数 + オレンジ写真
    media2, drawing_xml, drawing_rels, exts2 = _chosho_image_parts(orange_list, bd.get("photos_base64", {}))
    has_img = bool(media2)
    # 様式-1 全景写真(STEP5でチェックした1枚)
    zen_key = str(bd.get("zenkei_key", "") or "")
    zen_durl = (bd.get("photos_base64", {}) or {}).get(zen_key, "")
    media_z, drawing_z_xml, drawing_z_rels, ext_z = _chosho_zenkei_part(zen_durl)
    has_zen = bool(media_z)
    SHEET1RELS = "xl/worksheets/_rels/sheet2.xml.rels"
    ZEN_DRAW = "drawing3.xml"               # 様式-1 全景用の図面(drawing1=テンプレ既存,drawing2=様式-2)
    ZEN_RID = "rId2"                        # sheet2に追加する図面の関係ID(rId1=printerSettings)
    pages2 = max(1, (n_orange + 3) // 4)     # 1ページ4枚、最小1ページ
    rows2 = 44 + 38 * (pages2 - 1)           # 1=44,2=82,3=120,4=158,5=196
    # 印刷範囲内の全写真セル(説明文)を消去対象に(写真有無に関わらず)
    clear_refs = []
    for pair in range(pages2 * 2):           # 1ページ=2ペア(4枚)
        clear_refs.append("C%d" % (10 + 19 * pair))
        clear_refs.append("R%d" % (10 + 19 * pair))
    SHEET1 = "xl/worksheets/sheet2.xml"     # 様式-1
    SHEET2 = "xl/worksheets/sheet3.xml"     # 様式-2
    SHEET2RELS = "xl/worksheets/_rels/sheet3.xml.rels"
    CT = "[Content_Types].xml"
    WBOOK = "xl/workbook.xml"               # 様式-2の印刷範囲(Print_Area)を可変化
    DRAW_RID = "rId2"                       # sheet3に追加する図面の関係ID

    def esc(s):
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def patch_empty(xml, ref, val):
        v = esc(val)
        pat = re.compile(r'<c r="' + re.escape(ref) + r'"([^>]*?)/>')
        def _rep(m):
            return ('<c r="' + ref + '"' + m.group(1)
                    + ' t="inlineStr"><is><t xml:space="preserve">' + v + '</t></is></c>')
        return pat.subn(_rep, xml, count=1)[0]

    def patch_inline(xml, ref, val):
        v = esc(val)
        pat = re.compile(r'<c r="' + re.escape(ref) + r'"([^>]*?)(?:/>|>.*?</c>)', re.S)
        def _rep(m):
            attrs = re.sub(r'\s+t="[^"]*"', '', m.group(1))
            return ('<c r="' + ref + '"' + attrs
                    + ' t="inlineStr"><is><t xml:space="preserve">' + v + '</t></is></c>')
        return pat.subn(_rep, xml, count=1)[0]

    def patch_number(xml, ref, num):
        pat = re.compile(r'<c r="' + re.escape(ref) + r'"([^>]*?)(?:/>|>.*?</c>)', re.S)
        def _rep(m):
            attrs = re.sub(r'\s+t="[^"]*"', '', m.group(1))
            return '<c r="' + ref + '"' + attrs + '><v>' + str(num) + '</v></c>'
        return pat.subn(_rep, xml, count=1)[0]

    zin = zipfile.ZipFile(CHOSHO_TEMPLATE_PATH, "r")
    out_items = []
    for info in zin.infolist():
        data = zin.read(info.filename)
        if info.filename == SHEET1:
            xml = data.decode("utf-8")
            for ref, val in cells1.items():
                if val:
                    xml = patch_empty(xml, ref, val)
            # 全景写真あり: drawing参照を<extLst>直前へ挿入(スキーマ順守)
            if has_zen:
                pos = xml.rfind("<extLst")
                if pos != -1:
                    xml = xml[:pos] + ('<drawing r:id="%s"/>' % ZEN_RID) + xml[pos:]
                else:
                    xml = xml.replace("</worksheet>", '<drawing r:id="%s"/></worksheet>' % ZEN_RID, 1)
            data = xml.encode("utf-8")
        elif info.filename == SHEET2:
            xml = data.decode("utf-8")
            for ref, val in text2.items():
                if val:
                    xml = patch_inline(xml, ref, val)
            for ref, num in num2.items():
                xml = patch_number(xml, ref, num)
            # 手動改ページ(rowBreaks)をページ数に応じて設定(44,82,120,158)
            brk_rows = [44, 82, 120, 158][:pages2 - 1]
            if brk_rows:
                inner = "".join('<brk id="%d" max="29" man="1"/>' % b for b in brk_rows)
                new_rb = '<rowBreaks count="%d" manualBreakCount="%d">%s</rowBreaks>' % (len(brk_rows), len(brk_rows), inner)
            else:
                new_rb = '<rowBreaks count="0" manualBreakCount="0"/>'
            xml = re.sub(r'<rowBreaks\b.*?</rowBreaks>|<rowBreaks\b[^/]*/>', new_rb, xml, count=1, flags=re.S)
            # 印刷範囲内の写真セル説明文を消去(写真有無に関わらず)
            for cref in clear_refs:
                xml = patch_inline(xml, cref, "")
            # 写真画像あり: drawing参照を追加
            if has_img:
                xml = xml.replace("</colBreaks></worksheet>",
                                  '</colBreaks><drawing r:id="%s"/></worksheet>' % DRAW_RID, 1)
            data = xml.encode("utf-8")
        elif info.filename == SHEET2RELS and has_img:
            xml = data.decode("utf-8")
            rel = ('<Relationship Id="%s" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing2.xml"/>' % DRAW_RID)
            xml = xml.replace("</Relationships>", rel + "</Relationships>", 1)
            data = xml.encode("utf-8")
        elif info.filename == SHEET1RELS and has_zen:
            xml = data.decode("utf-8")
            relz = ('<Relationship Id="%s" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/%s"/>' % (ZEN_RID, ZEN_DRAW))
            xml = xml.replace("</Relationships>", relz + "</Relationships>", 1)
            data = xml.encode("utf-8")
        elif info.filename == CT and (has_img or has_zen):
            xml = data.decode("utf-8")
            adds = ""
            need_jpeg = ("jpeg" in exts2) or (ext_z == "jpeg")
            if need_jpeg and 'Extension="jpeg"' not in xml:
                adds += '<Default Extension="jpeg" ContentType="image/jpeg"/>'
            if has_img:
                adds += '<Override PartName="/xl/drawings/drawing2.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>'
            if has_zen:
                adds += '<Override PartName="/xl/drawings/%s" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>' % ZEN_DRAW
            xml = xml.replace("</Types>", adds + "</Types>", 1)
            data = xml.encode("utf-8")
        elif info.filename == WBOOK:
            xml = data.decode("utf-8")
            # 様式-2 の Print_Area の行数(...$AD$82 の82)を可変化
            xml = re.sub(r"(\u69d8\u5f0f-2'!\$A\$1:\$AD\$)\d+", lambda m: m.group(1) + str(rows2), xml, count=1)
            data = xml.encode("utf-8")
        out_items.append((info, data))
    zin.close()

    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, data in out_items:
            zout.writestr(info, data)
        if has_img:
            zout.writestr("xl/drawings/drawing2.xml", drawing_xml)
            zout.writestr("xl/drawings/_rels/drawing2.xml.rels", drawing_rels)
            for zipname, raw in media2:
                zout.writestr(zipname, raw)
        if has_zen:
            zout.writestr("xl/drawings/%s" % ZEN_DRAW, drawing_z_xml)
            zout.writestr("xl/drawings/_rels/%s.rels" % ZEN_DRAW, drawing_z_rels)
            for zipname, raw in media_z:
                zout.writestr(zipname, raw)


@app.post("/export-chosho")
async def export_chosho(request: Request):
    """道路橋記録様式(R6)を出力。様式-1にフリガナ・橋梁名・路線名を転記する。"""
    try:
        bd = await request.json()
    except Exception:
        bd = {}
    route_name = str(bd.get("路線名", "") or "")
    bridge_name = str(bd.get("橋梁名", "") or "")
    fname = "_".join([p for p in [route_name, bridge_name] if p]) or "道路橋記録様式"

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        out_path = tmp.name

    try:
        write_chosho(bd, out_path)
        return FileResponse(
            out_path,
            filename=f"{fname}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception:
        import traceback
        traceback.print_exc()
        if os.path.exists(out_path):
            os.unlink(out_path)
        raise HTTPException(status_code=500, detail=f"出力エラー: {traceback.format_exc()}")


@app.post("/export-data-format-csv")
async def export_data_format_csv(request: Request):
    """
    データ記録様式(csv)の出力。テンプレート未確定のため準備中。
    """
    raise HTTPException(status_code=501, detail="データ記録様式(csv)は準備中です")

CIRCLE_NUMS = '①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳㉑㉒㉓㉔㉕㉖'

def _dist(x1, y1, x2, y2):
    return math.sqrt((x1-x2)**2 + (y1-y2)**2)

def _get_insert(e):
    if e.dxftype() in ('TEXT', 'MTEXT'):
        return e.dxf.insert[0], e.dxf.insert[1]
    return None, None

def _get_text(e):
    if e.dxftype() == 'TEXT':
        return e.dxf.text.strip()
    elif e.dxftype() == 'MTEXT':
        return e.plain_text().strip()
    return ''

def _parse_damage_text(raw):
    """損傷テキスト1件を構造化データに変換"""
    # MTEXT末尾に埋め込まれた写真ラベル（写真-xx）を事前に抽出
    embedded_photo_label = None
    _pre = [l.strip() for l in raw.replace('\u3000', '').split('\n') if l.strip()]
    if _pre and re.match(r'^写真', _pre[-1]):
        embedded_photo_label = _pre[-1]
        _pre = _pre[:-1]
    raw = '\n'.join(_pre)
    # 改行分割後、損傷番号で始まらない行は直前行に結合
    raw_lines = [l.strip() for l in raw.replace('\u3000', '').split('\n') if l.strip()]
    lines = []
    in_extra = False
    for line in raw_lines:
        if line.startswith('※上記以外'):
            in_extra = True
            lines.append(line)
        elif in_extra:
            lines.append(line)
        elif not lines or re.match(r'^[' + CIRCLE_NUMS + r']', line) or line.startswith('※'):
            lines.append(line)
        else:
            lines[-1] = lines[-1] + line

    if not lines:
        return None

    result = {
        'member_name': None, 'symbol': None,
        'element_no_prev': None, 'element_no_current': None,
        'note': None, 'damages': [],
        'embedded_photo_label': embedded_photo_label
    }

    first = lines[0]

    # 備考行（※特に損傷なし → スキップ）
    if first == '※特に損傷なし':
        return None

    # ※上記以外の損傷 → 各行を「要素番号＋損傷の種類」で個別エントリとして解析
    if first.startswith('※'):
        member_dmg_pat = re.compile(
            r'^(.+?)\s+([A-Za-z]{2,4})([\d,→]+)\s*[:：]\s*(.+)$'
        )
        dmg_pat2 = re.compile(
            r'^([' + CIRCLE_NUMS + r'])(.+?)-([a-e])(?:→([a-e]))?(.*)' 
        )
        results = []
        for line in lines[1:]:
            if not line:
                continue
            m = member_dmg_pat.match(line)
            if m:
                member_name = m.group(1).strip()
                symbol      = m.group(2)
                elem_raw    = m.group(3)
                dmg_raw     = m.group(4)
                dm = dmg_pat2.match(dmg_raw)
                if dm:
                    parsed = _parse_detail(dm.group(5).strip())
                    ec = elem_raw.split('→')[1].strip() if '→' in elem_raw else elem_raw.strip()
                    ep = elem_raw.split('→')[0].strip() if '→' in elem_raw else None
                    dmg_entry = {
                        'damage_no':          dm.group(1),
                        'type':               dm.group(2),
                        'level_prev':         dm.group(3),
                        'level_current':      dm.group(4) or dm.group(3),
                        'detail':             parsed['quant'],
                        'pattern':            parsed['pattern'],
                        'bunrui':             parsed['bunrui'],
                        'bunrui_text':        parsed.get('bunrui_text', ''),
                    }
                    results.append({
                        'member_name':        member_name,
                        'symbol':             symbol,
                        'element_no_prev':    ep,
                        'element_no_current': ec,
                        'note':               None,
                        'is_extra':           True,
                        'damages':            [dmg_entry],
                    })
        return results if results else None

    # パターン1: 部材名変更あり「旧名 旧記号旧番→新名 新記号新番」
    m = re.match(r'^(.+?)\s+([A-Za-z]{2,4})(\d{4})→(.+?)\s+([A-Za-z]{2,4})(\d{4})\s*$', first)
    if m:
        result['member_name']        = m.group(4).strip()
        result['symbol']             = m.group(5)
        result['element_no_prev']    = m.group(3)
        result['element_no_current'] = m.group(6)
        result['is_arrow_current']   = True
    else:
        # パターン2: 要素番号変更あり「部材名 記号0101→0102,0103」
        m = re.match(r'^(.+?)\s+([A-Za-z]{2,4})(\d{4})→([\d,]+)', first)
        if m:
            result['member_name']        = m.group(1).strip()
            result['symbol']             = m.group(2)
            result['element_no_prev']    = m.group(3)
            result['element_no_current'] = m.group(4)
            result['is_arrow_current']   = True
        else:
            # パターン3: 要素番号のみ（単数・複数）
            m = re.match(r'^(.+?)\s+([A-Za-z]{2,4})([\d,]+)', first)
            if m:
                result['member_name']        = m.group(1).strip()
                result['symbol']             = m.group(2)
                result['element_no_current'] = m.group(3)
            else:
                result['member_name'] = first

    # 損傷行解析
    dmg_pat = re.compile(
        r'^([' + CIRCLE_NUMS + r'])'
        r'(.+?)-([a-e])'
        r'(?:→([a-e]))?'
        r'(.*)'
    )
    # 損傷種類変更パターン：⑰type-e→⑳type-e
    dmg_arrow_pat = re.compile(
        r'^([' + CIRCLE_NUMS + r'])(.+?)-([a-e])[^→]*→([' + CIRCLE_NUMS + r'])(.+?)-([a-e])(.*)'
    )
    for line in lines[1:]:
        # 損傷種類変更→矢印（⑰-e→⑳-e）を先にチェック
        ma = dmg_arrow_pat.match(line)
        if ma:
            parsed = _parse_detail(ma.group(7).strip())
            result['damages'].append({
                'damage_no':     ma.group(4),       # 今回の損傷番号
                'type':          ma.group(5).strip(), # 今回の損傷種類
                'level_prev':    ma.group(3),         # 前回程度
                'level_current': ma.group(6),         # 今回程度
                'type_prev':     ma.group(2).strip(), # 前回損傷種類（3-2用）
                'damage_no_prev': ma.group(1),
                'detail':        parsed['quant'],
                'pattern':       parsed['pattern'],
                'bunrui':        parsed['bunrui'],
                'bunrui_text':   parsed.get('bunrui_text', ''),
            })
            continue
        m = dmg_pat.match(line)
        if m:
            # →があれば前回→今回、なければ新規（level_prev=None）
            arrow = re.search(r'-([a-e])→([a-e])', line)
            if arrow:
                level_prev    = arrow.group(1)
                level_current = arrow.group(2)
            else:
                level_prev    = None
                level_current = m.group(3)

            parsed = _parse_detail(m.group(5).strip())
            result['damages'].append({
                'damage_no':     m.group(1),
                'type':          m.group(2).strip(),
                'level_prev':    level_prev,
                'level_current': level_current,
                'detail':        parsed['quant'],
                'pattern':       parsed['pattern'],
                'bunrui':        parsed['bunrui'],
                'bunrui_text':   parsed.get('bunrui_text', ''),
            })
        elif line and not line.startswith('写真'):
            result['damages'].append({'raw_line': line})

    # raw_lineのみ（※特に損傷なし等）はスキップ
    real_dmgs = [d for d in result.get('damages', []) if 'damage_no' in d or 'type' in d]
    if not real_dmgs and not result.get('element_no_current'):
        return None
    result['damages'] = real_dmgs if real_dmgs else result.get('damages', [])
    # 要素番号がカンマ区切りの場合も1エントリのまま保持（カンマ区切りで保持）
    return result


def _parse_detail(raw_detail: str) -> dict:
    """
    損傷行の末尾部分から定量値・分類・パターン・対策工区分を分離する。
    例: '(大大)[5.0mm/0.5m未満2.5mm]パターン②(Ⅱ)'
      → quant='5.0mm/0.5m未満2.5mm', pattern='②', bunrui='', taisaku='(Ⅱ)'
    例: '(欠損)(I)'
      → quant='', pattern='', bunrui='', taisaku='(I)'
    """
    if not raw_detail:
        return {'quant': '', 'pattern': '', 'bunrui': '', 'taisaku': ''}

    s = raw_detail.strip()

    # パターン番号を抽出: (パターン①)/(ﾊﾟﾀｰﾝ①)/パターン①/ﾊﾟﾀｰﾝ① いずれにも対応
    pattern = ''
    m = re.search(r'[（\(](?:パターン|ﾊﾟﾀｰﾝ)([^）\)]+)[）\)]', s)
    if m:
        pattern = m.group(1).strip()
        s = s[:m.start()] + s[m.end():]
    else:
        m = re.search(r'(?:パターン|ﾊﾟﾀｰﾝ)([①-⑳㉑-㉛])', s)
        if m:
            pattern = m.group(1)
            s = s[:m.start()] + s[m.end():]

    # 分類を抽出: (分類6:xxx) or (分類1)
    bunrui = ''
    bunrui_text = ''
    m = re.search(r'[（\(]分類(\d)(?::([^）\)]+))?[）\)]', s)
    if m:
        bunrui = m.group(1)
        bunrui_text = m.group(2).strip() if m.group(2) else ''
        s = s[:m.start()] + s[m.end():]

    # []の中身=定量的に取得した値（taisaku除去より先に行う）
    quant = ''
    m = re.search(r'\[([^\]]+)\]', s)
    if m:
        quant = m.group(1).strip()
        s = s[:m.start()] + s[m.end():]

    # 対策工区分を抽出: (I)(II)(Ⅰ)(Ⅱ)等 → 保持するが定量値からは除く
    taisaku = ''
    m = re.search(r'[（\(][IVXⅠⅡⅢⅣⅤⅰⅱⅲⅳⅴ]+[）\)]$', s.strip())
    if m:
        taisaku = m.group(0)
        s = s[:m.start()].strip()

    return {
        'quant': quant,
        'pattern': pattern,
        'bunrui': bunrui,
        'bunrui_text': bunrui_text,
        'taisaku': taisaku,
        'detail_raw': s.strip(),
    }


def _clean_detail(detail: str) -> str:
    """後方互換用：定量値のみを返す"""
    return _parse_detail(detail).get('quant', '')


def extract_dxf_spans(filepath: str, direction: str = 'col') -> list:
    """
    DXFファイルから径間リストを返す。
    各径間の視点ブロック内のM-STR-HTXTから部材リストも生成する。
    """
    import ezdxf
    doc = ezdxf.readfile(filepath)
    msp = doc.modelspace()

    # D-TTL-LINEの枠を収集
    boxes = []
    for e in msp:
        if e.dxf.layer == 'D-TTL-LINE' and e.dxftype() == 'LWPOLYLINE':
            pts = list(e.get_points())
            xs = [p[0] for p in pts]
            ys = [p[1] for p in pts]
            boxes.append({
                'xmin': min(xs), 'xmax': max(xs),
                'ymin': min(ys), 'ymax': max(ys),
                'title': None, 'span_no': 1, 'members': []
            })

    # 視点タイトルを割り当て
    for e in msp:
        if e.dxf.layer == '視点タイトル' and e.dxftype() == 'TEXT':
            x, y = e.dxf.insert[0], e.dxf.insert[1]
            for b in boxes:
                if b['xmin'] <= x <= b['xmax'] and b['ymin'] <= y <= b['ymax']:
                    b['title'] = e.dxf.text
                    break

    boxes = [b for b in boxes if b['title']]
    if not boxes:
        return []

    # HTMLと同じアルゴリズムで径間グループ化
    avg_size = sum(b['xmax'] - b['xmin'] for b in boxes) / len(boxes)
    threshold = avg_size * 0.5

    if direction == 'col':
        coords = sorted(set(b['xmin'] for b in boxes))
        groups = [[coords[0]]]
        for c in coords[1:]:
            if c - groups[-1][0] > threshold:
                groups.append([c])
            else:
                groups[-1].append(c)
        for b in boxes:
            idx = next(
                (i for i, g in enumerate(groups) if any(abs(c - b['xmin']) <= threshold for c in g)), 0
            )
            b['span_no'] = idx + 1
        boxes.sort(key=lambda b: (b['span_no'], -b['ymax']))
    else:
        coords = sorted(set(b['ymax'] for b in boxes), reverse=True)
        groups = [[coords[0]]]
        for c in coords[1:]:
            if groups[-1][0] - c > threshold:
                groups.append([c])
            else:
                groups[-1].append(c)
        for b in boxes:
            idx = next(
                (i for i, g in enumerate(groups) if any(abs(c - b['ymax']) <= threshold for c in g)), 0
            )
            b['span_no'] = idx + 1
        boxes.sort(key=lambda b: (-b['ymax'], b['span_no']))

    # M-STR-HTXTから部材を抽出（TEXT/MTEXTのみ、LEADER除外）
    member_pat = re.compile(r'^(.+?)\s+([A-Z][a-z]+)(\d{4})')
    seen = set()  # (span_no, symbol, elem_no) 重複除去

    for e in msp:
        if e.dxftype() not in ('TEXT', 'MTEXT'):
            continue
        try:
            x, y = e.dxf.insert[0], e.dxf.insert[1]
            text = e.dxf.text if e.dxftype() == 'TEXT' else e.plain_text()
        except Exception:
            continue
        if not text or text.startswith('写真') or re.match(r'^[\d.]+$', text):
            continue
        # 全行走査（※上記以外の損傷ブロック等、2行目以降の部材行も拾う）
        for ln in text.split('\n'):
            ln = ln.strip()
            if not ln or ln.startswith('写真') or ln.startswith('※') or re.match(r'^[①-⑳㉑-㉛]', ln):
                continue
            m = member_pat.match(ln)
            if not m:
                continue
            name = m.group(1).strip()
            symbol = m.group(2)
            elem_nos = [m.group(3)]
            is_arrow = False
            # →があれば変更後の要素番号・記号を採用（カンマ区切りは全て展開）
            arrow = re.search(r'→(?:[^\s]+\s+)?([A-Z][a-z]+)?(\d{4}(?:,\d{4})*)', ln)
            if arrow and arrow.group(2):
                elem_nos = arrow.group(2).split(',')
                is_arrow = True
                if arrow.group(1):
                    symbol = arrow.group(1)
            for b in boxes:
                if b['xmin'] <= x <= b['xmax'] and b['ymin'] <= y <= b['ymax']:
                    for elem_no in elem_nos:
                        key = (b['span_no'], symbol, elem_no)
                        if key not in seen:
                            seen.add(key)
                            entry = {
                                'koshu': '', 'zairyo': '',
                                'name': name, 'symbol': symbol, 'element_no': elem_no
                            }
                            if is_arrow:
                                entry['is_arrow_current'] = True
                            b['members'].append(entry)
                    break

    # 径間ごとにまとめる
    span_map = {}
    for b in boxes:
        sp = b['span_no']
        if sp not in span_map:
            span_map[sp] = {'views': [], 'members': []}
        span_map[sp]['views'].append(b['title'])
        span_map[sp]['members'].extend(b['members'])

    return [
        {'span_no': sp, 'views': span_map[sp]['views'], 'members': span_map[sp]['members']}
        for sp in sorted(span_map.keys())
    ]


def extract_element_numbers(filepath: str) -> dict:
    """
    DXFファイルから部材記号と要素番号の対応を抽出する。
    M-STR-HTXT レイヤーのテキストから「部材記号+4桁数字」パターンを収集。
    戻り値: {"Mg": ["0101","0102",...], "Cr": [...], ...}
    """
    import ezdxf
    doc = ezdxf.readfile(filepath)
    msp = doc.modelspace()

    # 部材記号パターン: 2〜4英字 + 4桁数字
    elem_pat = re.compile(r'([A-Za-z]{2,4})(\d{4})')
    result = {}

    for e in msp:
        if e.dxftype() not in ('TEXT', 'MTEXT'):
            continue
        try:
            if e.dxftype() == 'TEXT':
                text = e.dxf.text
            else:
                text = e.plain_text()
        except Exception:
            continue
        if not text:
            continue
        # テキスト内の全パターンを収集
        for m in elem_pat.finditer(text):
            symbol = m.group(1)
            elem_no = m.group(2)
            if symbol not in result:
                result[symbol] = []
            if elem_no not in result[symbol]:
                result[symbol].append(elem_no)

    # 各部材の要素番号をソート
    for symbol in result:
        result[symbol].sort()

    return result


def parse_dxf_file(filepath: str, direction: str = 'col') -> list:
    import ezdxf
    """
    DXFファイルを解析してグループ別損傷リストを返す
    direction: 'col'=列方向(x→y順) 'row'=行方向(y→x順)
    """
    doc = ezdxf.readfile(filepath)
    msp = doc.modelspace()

    # D-TTL-LINEの矩形グループ
    boxes = []
    for e in msp:
        if e.dxf.layer == 'D-TTL-LINE' and e.dxftype() == 'LWPOLYLINE':
            pts = list(e.get_points())
            xs = [p[0] for p in pts]
            ys = [p[1] for p in pts]
            boxes.append({
                'xmin': min(xs), 'xmax': max(xs),
                'ymin': min(ys), 'ymax': max(ys),
                'title': None, 'damages': []
            })

    def in_box(b, x, y):
        return b['xmin'] <= x <= b['xmax'] and b['ymin'] <= y <= b['ymax']

    # 視点タイトルをグループに割り当て
    for e in msp:
        if e.dxf.layer == '視点タイトル' and e.dxftype() == 'TEXT':
            x, y = _get_insert(e)
            for b in boxes:
                if in_box(b, x, y):
                    b['title'] = e.dxf.text
                    break

    # テキスト収集（レイヤー指定なし・ボックス内のみ対象）
    # 除外レイヤー：別処理のもののみ
    EXCLUDE_LAYERS = {'視点タイトル', 'D-TTL-LINE'}
    damage_texts, photo_labels, photo_files = [], [], []
    for e in msp:
        if e.dxftype() not in ('TEXT', 'MTEXT'):
            continue
        if e.dxf.layer in EXCLUDE_LAYERS:
            continue
        x, y = _get_insert(e)
        if x is None:
            continue
        # ボックス外は無視
        if not any(in_box(b, x, y) for b in boxes):
            continue
        text = _get_text(e)
        if not text:
            continue

        # 写真番号レイヤーは従来通り別処理
        if e.dxf.layer == '写真番号':
            if re.match(r'^[A-Za-z_\d]', text):
                for part in [t.strip() for t in text.split(',') if t.strip()]:
                    photo_files.append({'x': x, 'y': y, 'text': part})
            continue

        # 数値のみ・XY座標記号はスキップ
        if re.match(r'^[\d.]+$', text) or re.match(r'^[XY]\d+$', text):
            continue

        # 独立した写真ラベル（「写真」で始まり改行なし）
        if re.match(r'^写真', text) and '\n' not in text:
            h = e.dxf.height if e.dxf.hasattr('height') else 7.0
            photo_labels.append({'x': x, 'y': y, 'text': text, 'h': h})
            continue

        # 損傷テキスト（MTEXT末尾の写真ラベル埋め込みはparse時に抽出）
        damage_texts.append({'x': x, 'y': y, 'text': text})

    # 損傷テキストをグループに割り当て（座標付きで保存）
    for dt in damage_texts:
        for b in boxes:
            if in_box(b, dt['x'], dt['y']):
                b['damages'].append({
                    'x': dt['x'], 'y': dt['y'], 'raw': dt['text'],
                    'photo_label': None, 'photo_file': None
                })
                break

    # 各ボックス内でX座標が近い行を結合（部材行+損傷行を1ブロックに）
    CIRCLE_PAT = re.compile(r'^[①-⑳㉑-㉛※]')
    X_THRESH = 20  # X座標の許容誤差

    for b in boxes:
        if not b['damages']:
            continue

        # TEXTによる※上記以外箇条書き群を1rawに束ねる
        items_raw = sorted(b['damages'], key=lambda d: (-d['y'], d['x']))
        bundled = []
        skip_idxs = set()
        for i, item in enumerate(items_raw):
            if i in skip_idxs:
                continue
            if item['raw'].strip() == '※上記以外の損傷':
                combined = item['raw']
                for k in range(len(items_raw)):
                    if k == i or k in skip_idxs:
                        continue
                    other = items_raw[k]
                    if (other['y'] < item['y'] and
                            abs(other['x'] - item['x']) <= X_THRESH * 3 and
                            item['y'] - other['y'] <= 70):
                        combined = combined + '\n' + other['raw']
                        skip_idxs.add(k)
                bundled.append({'x': item['x'], 'y': item['y'], 'raw': combined,
                                'photo_label': None, 'photo_file': None})
            else:
                bundled.append(item)
        b['damages'] = bundled

        # X座標でグループ化（X_THRESH以内を同列とみなす）
        items = sorted(b['damages'], key=lambda d: d['x'])
        x_groups = []
        for item in items:
            placed = False
            for grp in x_groups:
                if abs(item['x'] - grp[0]['x']) <= X_THRESH:
                    grp.append(item)
                    placed = True
                    break
            if not placed:
                x_groups.append([item])

        # 各X列内でY降順（上→下）に並べて部材行+損傷行を結合
        merged = []
        for grp in x_groups:
            grp.sort(key=lambda d: -d['y'])  # Y降順（上が先）
            for item in grp:
                is_extra = item['raw'].startswith('※上記以外')
                if not is_extra and CIRCLE_PAT.match(item['raw']) and merged and abs(item['x'] - merged[-1]['x']) <= X_THRESH:
                    merged[-1]['raw'] = merged[-1]['raw'] + '\n' + item['raw']
                else:
                    merged.append({'x': item['x'], 'y': item['y'], 'raw': item['raw'],
                                   'photo_label': None, 'photo_file': None})
        b['damages'] = merged

    # 2段階マッピング：旗揚げ構造（損傷→写真ラベル→P番号）を位置関係で紐付け

    # ボックスごとの相対閾値を計算
    for b in boxes:
        b['h'] = b['ymax'] - b['ymin']
        b['w'] = b['xmax'] - b['xmin']

    # STEP1: 写真ラベルとP番号を旗揚げペアとしてマッチング
    flag_pairs = []
    used_pf = set()

    for pl in photo_labels:
        for b in boxes:
            if not in_box(b, pl['x'], pl['y']):
                continue
            y_thresh = b['h'] * 0.08
            x_thresh = pl.get('h', 7.0) * 15
            matched_pfs = []
            for idx, pf in enumerate(photo_files):
                if idx in used_pf:
                    continue
                if not in_box(b, pf['x'], pf['y']):
                    continue
                if abs(pf['y'] - pl['y']) <= y_thresh and abs(pf['x'] - pl['x']) <= x_thresh:
                    matched_pfs.append((idx, pf))
            if matched_pfs:
                matched_pfs.sort(key=lambda t: abs(t[1]['x'] - pl['x']))
                for idx, pf in matched_pfs:
                    used_pf.add(idx)
                flag_pairs.append({'label': pl, 'files': [pf for _, pf in matched_pfs], 'box': b})
            break

    # STEP2: 旗揚げペアを損傷に紐付け（同X列かつ直上の損傷）
    for pair in flag_pairs:
        pl = pair['label']
        b = pair['box']
        x_thresh = b['w'] * 0.1
        best, best_d = None, 1e9
        for d in b['damages']:
            if abs(d['x'] - pl['x']) > x_thresh:
                continue
            if d['y'] < pl['y']:
                continue
            dist = d['y'] - pl['y']
            if dist < best_d:
                best_d, best = dist, d
        if best:
            label_text = pair['label']['text']
            best['photo_label'] = (best['photo_label'] + ',' + label_text) if best['photo_label'] else label_text
            file_texts = ','.join(pf['text'] for pf in pair['files'])
            best['photo_file'] = (best['photo_file'] + ',' + file_texts) if best['photo_file'] else file_texts

    # STEP3: 損傷ごとに最近傍の未使用P番号を割り当て（距離閾値内のみ）
    # 埋め込み写真ラベルを持つ損傷を優先してマッチング
    used_pf2 = set(used_pf)
    unmatched_pf = [(idx, pf) for idx, pf in enumerate(photo_files) if idx not in used_pf2]
    for b in boxes:
        box_w = b['xmax'] - b['xmin']
        thresh = box_w * 0.15  # ボックス幅の15%以内
        # 埋め込みラベルありを先に処理
        ordered = sorted(b['damages'], key=lambda d: (0 if d.get('photo_label') else 1))
        for d in ordered:
            if d['photo_file']:
                continue  # STEP1/2で解決済み
            # 損傷番号（①〜㉛）を含まない非損傷テキストはスキップ
            if not re.search(r'[①-⑳㉑-㉛]', d.get('raw', '')):
                continue
            best_idx, best_pf, best_d = None, None, 1e9
            for idx, pf in unmatched_pf:
                if not in_box(b, pf['x'], pf['y']):
                    continue
                dist = _dist(pf['x'], pf['y'], d['x'], d['y'])
                if dist < best_d and dist <= thresh:
                    best_d, best_idx, best_pf = dist, idx, pf
            if best_pf:
                d['photo_file'] = best_pf['text']
                unmatched_pf = [(i, p) for i, p in unmatched_pf if i != best_idx]

    # ソート
    if direction == 'col':
        boxes.sort(key=lambda b: (b['xmin'], -b['ymax']))
    else:
        boxes.sort(key=lambda b: (-b['ymax'], b['xmin']))

    # 結果整形
    groups = []
    for b in boxes:
        parsed = []
        for d in b['damages']:
            raw = d['raw']
            p = _parse_damage_text(raw)
            if not p:
                continue
            # ※上記以外の損傷はリストで返る → 1旗揚げ=1エントリに統合
            if isinstance(p, list):
                if not p:
                    continue
                merged_entry = p[0]
                for extra_item in p[1:]:
                    for dmg in extra_item.get('damages', []):
                        dmg['member_name']        = extra_item.get('member_name', '')
                        dmg['symbol']             = extra_item.get('symbol', '')
                        dmg['element_no_current'] = extra_item.get('element_no_current', '')
                        merged_entry['damages'].append(dmg)
                merged_entry['photo_label'] = d['photo_label']
                merged_entry['photo_file']  = d['photo_file']
                parsed.append(merged_entry)
            else:
                p['photo_label'] = d['photo_label'] or p.pop('embedded_photo_label', None)
                p['photo_file']  = d['photo_file']
                parsed.append(p)
        groups.append({'title': b['title'], 'damages': parsed})

    return groups


def extract_render_data(filepath: str, direction: str = 'col') -> list:
    """各視点ボックスのSVG描画用データを返す"""
    import ezdxf
    doc = ezdxf.readfile(filepath)
    msp = doc.modelspace()

    DRAW_LAYERS = {'構造物', '損傷スケッチ', 'M-STR-HTXT', '写真番号', 'D-TTL-LINE', '視点タイトル', 'D-STR', 'D-STR-STR1'}

    # D-TTL-LINE枠収集
    boxes = []
    for e in msp:
        if e.dxf.layer == 'D-TTL-LINE' and e.dxftype() == 'LWPOLYLINE':
            pts = list(e.get_points())
            xs = [p[0] for p in pts]; ys = [p[1] for p in pts]
            boxes.append({'xmin': min(xs), 'xmax': max(xs), 'ymin': min(ys), 'ymax': max(ys),
                          'title': None, 'span_no': 1, 'entities': []})

    # 視点タイトル割当
    for e in msp:
        if e.dxf.layer == '視点タイトル' and e.dxftype() == 'TEXT':
            x, y = e.dxf.insert[0], e.dxf.insert[1]
            for b in boxes:
                if b['xmin'] <= x <= b['xmax'] and b['ymin'] <= y <= b['ymax']:
                    b['title'] = e.dxf.text; break
    boxes = [b for b in boxes if b['title']]
    if not boxes:
        return []

    # 径間グループ化（extract_dxf_spansと同じロジック）
    avg_size = sum(b['xmax'] - b['xmin'] for b in boxes) / len(boxes)
    threshold = avg_size * 0.5
    if direction == 'col':
        coords = sorted(set(b['xmin'] for b in boxes))
        groups = [[coords[0]]]
        for c in coords[1:]:
            if c - groups[-1][0] > threshold: groups.append([c])
            else: groups[-1].append(c)
        for b in boxes:
            idx = next((i for i, g in enumerate(groups) if any(abs(c - b['xmin']) <= threshold for c in g)), 0)
            b['span_no'] = idx + 1
        boxes.sort(key=lambda b: (b['span_no'], -b['ymax']))
    else:
        coords = sorted(set(b['ymax'] for b in boxes), reverse=True)
        groups = [[coords[0]]]
        for c in coords[1:]:
            if groups[-1][0] - c > threshold: groups.append([c])
            else: groups[-1].append(c)
        for b in boxes:
            idx = next((i for i, g in enumerate(groups) if any(abs(c - b['ymax']) <= threshold for c in g)), 0)
            b['span_no'] = idx + 1
        boxes.sort(key=lambda b: (-b['ymax'], b['span_no']))

    def in_box(b, x, y, margin=50):
        return b['xmin']-margin <= x <= b['xmax']+margin and b['ymin']-margin <= y <= b['ymax']+margin

    # エンティティ収集
    # 構造線（LINE/POLYLINE/CIRCLE）：レイヤー指定で制御
    # 注記（TEXT/MTEXT/LEADER）：レイヤー指定なし・ボックス内フィルタのみ
    ANNOTATION_EXCLUDE = {'D-TTL-LINE', '視点タイトル', 'Defpoints'}
    for e in msp:
        layer = e.dxf.layer
        etype = e.dxftype()
        ent = None

        if etype == 'LINE':
            if layer not in DRAW_LAYERS: continue
            s, en = e.dxf.start, e.dxf.end
            ent = {'type': 'line', 'layer': layer,
                   'x1': float(s[0]), 'y1': float(s[1]),
                   'x2': float(en[0]), 'y2': float(en[1])}
        elif etype == 'LWPOLYLINE':
            if layer not in DRAW_LAYERS: continue
            pts = list(e.get_points())
            coords = [[float(p[0]), float(p[1])] for p in pts]
            closed = bool(e.dxf.flags & 1)
            ent = {'type': 'polyline', 'layer': layer, 'points': coords, 'closed': closed}
        elif etype in ('TEXT', 'MTEXT'):
            if layer in ANNOTATION_EXCLUDE: continue
            x, y = _get_insert(e)
            if x is None: continue
            text = _get_text(e)
            if not text: continue
            try: h = float(e.dxf.height)
            except:
                try: h = float(e.dxf.char_height)
                except: h = 7.0
            ent = {'type': 'text', 'layer': layer, 'x': float(x), 'y': float(y), 'text': text, 'height': h}
        elif etype == 'LEADER':
            if layer in ANNOTATION_EXCLUDE: continue
            pts = list(e.vertices)
            coords = [[float(p[0]), float(p[1])] for p in pts]
            ent = {'type': 'leader', 'layer': layer, 'points': coords}
        elif etype == 'CIRCLE':
            if layer not in DRAW_LAYERS: continue
            c = e.dxf.center
            ent = {'type': 'circle', 'layer': layer,
                   'cx': float(c[0]), 'cy': float(c[1]), 'r': float(e.dxf.radius)}

        if ent:
            # エンティティの代表座標を取得してボックスに割り当て
            if etype == 'LINE':
                mx, my = (ent['x1']+ent['x2'])/2, (ent['y1']+ent['y2'])/2
            elif etype == 'LWPOLYLINE':
                mx = sum(p[0] for p in ent['points']) / len(ent['points'])
                my = sum(p[1] for p in ent['points']) / len(ent['points'])
            elif etype in ('TEXT', 'MTEXT'):
                mx, my = ent['x'], ent['y']
            elif etype == 'LEADER':
                mx, my = ent['points'][0][0], ent['points'][0][1]
            elif etype == 'CIRCLE':
                mx, my = ent['cx'], ent['cy']
            else:
                continue
            for b in boxes:
                if in_box(b, mx, my):
                    b['entities'].append(ent)
                    break

    # 結果整形
    result = []
    for b in boxes:
        result.append({
            'title': b['title'],
            'span_no': b['span_no'],
            'xmin': b['xmin'], 'xmax': b['xmax'],
            'ymin': b['ymin'], 'ymax': b['ymax'],
            'entities': b['entities']
        })
    return result


@app.post("/extract-dxf")
async def extract_dxf(
    file: UploadFile = File(...),
    direction: str = Form("col")
):
    """
    DXFファイルを解析してグループ別損傷リストを返す
    direction: 'col'=列方向 / 'row'=行方向
    """
    suffix = os.path.splitext(file.filename)[1].lower()
    if suffix != ".dxf":
        raise HTTPException(status_code=400, detail=".dxfファイルを指定してください")

    with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        groups = parse_dxf_file(tmp_path, direction)
        element_numbers = extract_element_numbers(tmp_path)
        dxf_spans = extract_dxf_spans(tmp_path, direction)
        render_data = extract_render_data(tmp_path, direction)
        with open(tmp_path, "r", encoding="utf-8", errors="replace") as f:
            raw_text = f.read()
        return {"status": "ok", "groups": groups, "element_numbers": element_numbers, "dxf_spans": dxf_spans, "render_data": render_data, "raw_text": raw_text}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DXF解析エラー: {str(e)}")
    finally:
        os.unlink(tmp_path)
